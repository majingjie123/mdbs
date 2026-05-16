import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import re
import os
from datetime import datetime

# 预编译正则：剔除 ASCII 控制字符，避免每次调用 _clean_value 都重新编译
_CLEAN_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
from models.db_storage import DBStorage
from core.db_operations import DBOperations
from ui.progress_dialog import ProgressDialog
from core.sql_completer import SQLCompleter, CompletionItem
from core.metadata_cache import MetadataCache
from core.scratch_manager import ScratchManager
from core.theme import get_theme_colors, is_dark_theme
from tkinter import simpledialog

class SQLWorkbench(ttk.Frame):
    def __init__(self, parent, conn_id, db_name=None, schema_name=None, on_save_callback=None, load_draft=True, cancel_event=None, script_path=None, initial_sql=None):
        super().__init__(parent)
        self.conn_id = conn_id
        self.storage = DBStorage()
        self.conn_data = self.storage.get_connection(conn_id)
        if not self.conn_data:
            messagebox.showerror("错误", "无法加载连接信息")
            return
            
        self.db_name = db_name or self.conn_data.get('database', 'postgres' if self.conn_data.get('db_type') == 'PostgreSQL' else '')
        self.schema_name = schema_name or ('public' if self.conn_data.get('db_type') == 'PostgreSQL' else None)
        self.on_save_callback = on_save_callback
        self.db_ops = DBOperations()
        self.settings = self.storage.get_settings()
        self.cancel_event = cancel_event
        
        self.initial_sql = initial_sql
        self.current_query_table = None
        self.current_query_pks = []
        self.pending_updates = {}
        self.original_values = {}
        self.all_rows_data = []
        
        self.completer = SQLCompleter(conn_id, conn_data=self.conn_data, db_ops=self.db_ops)
        self.metadata_cache = MetadataCache()
        self._current_db = self.db_name  # 补全用的当前数据库，切换库时同步更新
        
        self.comp_window = None
        self.comp_listbox = None
        self.scratch_manager = ScratchManager()
        self.is_running = False
        
        # 分页相关状态
        self.page_size = tk.IntVar(value=5000)
        self.current_page = 1
        self.total_pages = 1
        self.all_cols = []
        self.all_rows = []
        
        # 排序与筛选状态
        self.sort_column = None      # 当前排序列名
        self.sort_ascending = True   # 升序/降序
        self.filter_column = None    # 筛选列名
        self.filter_value = ""       # 筛选值
        self.original_all_rows = []  # 原始数据备份（用于清除筛选）
        
        # 绑定脚本路径
        self.autosave_id = f"{self.conn_data['name']}_{self.db_name}"
        self.script_path = script_path or os.path.join(self.scratch_manager._base_dir, f"autosave_{self.autosave_id}.sql")
        self._autosave_debounce_id = None  # 防抖定时器 ID
        self._last_autosave_content = ""    # 上次保存内容快照，避免冗余保存
        
        self._init_ui()
        self._apply_theme()
        
        # 加载持久化日志 (按路径)
        self._load_persistent_logs()
        
        if load_draft:
            if script_path:
                content = self.scratch_manager.get_script_content_by_path(script_path)
            else:
                content = self.scratch_manager.load_autosave(self.autosave_id)
                
            if content:
                self.editor.insert("1.0", content)
                self._highlight_syntax()
        
        self._start_autosave_timer()
        self._refresh_metadata()
        self._load_database_list()
        self._load_schema_list()

        # initial_sql 自动执行（用于双击表直接浏览数据）
        if self.initial_sql:
            self.after(150, lambda: self._execute_initial_sql())

    def _execute_initial_sql(self):
        """插入初始 SQL 并自动执行（用于双击表浏览数据）"""
        if not self.winfo_exists() or not self.initial_sql:
            return
        self.editor.insert("1.0", self.initial_sql)
        self._highlight_syntax()
        self.initial_sql = None  # 避免重复执行
        self._run_query()

    def _refresh_metadata(self):
        """刷新当前库/模式的元数据缓存"""
        schema = self.schema_name if self.conn_data.get('db_type') == 'PostgreSQL' else None
        self.metadata_cache.fetch_tables_for_db(
            self.conn_data['id'], self.conn_data, self.db_name, self.db_ops, 
            schema=schema,
            callback=lambda: self._log_msg(f"元数据已更新: {self.db_name}.{schema if schema else ''}")
        )
        self._log_msg(f"正在后台刷新元数据: {self.db_name}...")

    def _load_database_list(self):
        """异步加载数据库列表并填充选择器"""
        def run():
            try:
                dbs = self.db_ops.get_databases(self.conn_data)
                self.after(0, lambda: self.db_selector.configure(values=dbs))
            except: pass
        threading.Thread(target=run, daemon=True).start()

    def _load_schema_list(self):
        """异步加载指定库下的 Schema (仅 PG)"""
        if self.conn_data.get('db_type') != 'PostgreSQL': return
        if not hasattr(self, 'schema_selector'): return
        def run():
            try:
                schemas = self.db_ops.get_schemas(self.conn_data, database=self.db_name)
                def update_ui():
                    self.schema_selector.configure(values=schemas)
                    # 确保当前值在列表中，否则回退到 public
                    current = self.schema_selector.get()
                    if current not in schemas:
                        fallback = 'public' if 'public' in schemas else (schemas[0] if schemas else '')
                        self.schema_selector.set(fallback)
                        self.schema_name = fallback
                self.after(0, update_ui)
            except: pass
        threading.Thread(target=run, daemon=True).start()

    def _on_db_changed(self, event):
        """用户手动切换库"""
        new_db = self.db_selector.get()
        if new_db == self.db_name: return
        
        self.db_name = new_db
        self._current_db = new_db  # 同步更新补全用的数据库
        self._log_msg(f"已切换数据库上下文至: {new_db}", "INFO")
        
        # PostgreSQL: 切换数据库后重置模式为 public 并刷新列表
        if self.conn_data.get('db_type') == 'PostgreSQL':
            self.schema_name = 'public'
            if hasattr(self, 'schema_selector'):
                self.schema_selector.set('public')
            self._load_schema_list()
        
        self._refresh_metadata()

        # 更新自动保存 ID
        self.autosave_id = f"{self.conn_data['name']}_{self.db_name}"

        # 同步 AI 聊天面板的数据库信息
        if hasattr(self, 'chat_panel'):
            self.chat_panel.db_name = new_db

    def _on_schema_changed(self, event):
        """用户手动切换模式"""
        new_schema = self.schema_selector.get()
        if new_schema == self.schema_name: return
        self.schema_name = new_schema
        self._log_msg(f"已切换模式上下文至: {new_schema}", "INFO")
        self._refresh_metadata()

    def _on_editor_modified(self, event=None):
        """编辑器内容变更时触发：防抖 5 秒后自动保存"""
        try:
            self.editor.tk.call(self.editor._w, "edit", "modified", 0)
        except tk.TclError:
            return
        # 重置防抖定时器
        if self._autosave_debounce_id:
            self.after_cancel(self._autosave_debounce_id)
        self._autosave_debounce_id = self.after(5000, self._debounced_autosave)

    def _debounced_autosave(self):
        """防抖到期后执行自动保存"""
        self._autosave_debounce_id = None
        self._do_autosave()

    def _do_autosave(self):
        """执行实际自动保存（仅草稿模式、内容有变化时）"""
        if "autosave_" not in os.path.basename(self.script_path):
            return
        try:
            if not self.winfo_exists():
                return
        except:
            return
        content = self.editor.get("1.0", tk.END).strip()
        if not content or content == self._last_autosave_content:
            return
        self._last_autosave_content = content
        def run_save(c):
            try: self.scratch_manager.save_autosave(self.autosave_id, c)
            except: pass
        threading.Thread(target=run_save, args=(content,), daemon=True).start()

    def _start_autosave_timer(self):
        """后备定时器（每 30 秒检测一次，仅在事件驱动遗漏时兜底）"""
        try:
            if not self.winfo_exists():
                return
            self._do_autosave()
        except:
            pass
        self.after(30000, self._start_autosave_timer)

    def _init_ui(self):
        self.toolbar = tk.Frame(self, padx=5, pady=5)
        self.toolbar.pack(side="top", fill="x")
        
        self.exec_btn = ttk.Button(self.toolbar, text="⚡ 执行 (F5)", command=lambda: self._run_query(needs_confirm=False))
        self.exec_btn.pack(side="left", padx=2)
        
        self.stop_btn = ttk.Button(self.toolbar, text="🛑 停止", command=self._stop_query, state="disabled")
        self.stop_btn.pack(side="left", padx=2)

        ttk.Button(self.toolbar, text="💾 保存 (Ctrl+S)", command=self._save_script).pack(side="left", padx=2)
        ttk.Button(self.toolbar, text="🪄 美化 SQL", command=self._format_sql).pack(side="left", padx=2)
        ttk.Button(self.toolbar, text="📦 执行 SQL 文件", command=self._run_sql_file).pack(side="left", padx=2)
        
        ttk.Separator(self.toolbar, orient="vertical").pack(side="left", padx=10, fill="y")
        
        ttk.Label(self.toolbar, text="🗄️ 数据库:").pack(side="left", padx=2)
        self.db_selector = ttk.Combobox(self.toolbar, width=20, state="readonly")
        self.db_selector.pack(side="left", padx=5)
        self.db_selector.set(self.db_name)
        self.db_selector.bind("<<ComboboxSelected>>", self._on_db_changed)
        
        # PostgreSQL 模式选择器 (仅 PG 连接时显示)
        self.schema_frame = tk.Frame(self.toolbar)
        if self.conn_data.get('db_type') == 'PostgreSQL':
            self.schema_frame.pack(side="left", padx=2)
            ttk.Label(self.schema_frame, text="🏺 模式:").pack(side="left", padx=2)
            self.schema_selector = ttk.Combobox(self.schema_frame, width=18, state="readonly")
            self.schema_selector.pack(side="left", padx=5)
            self.schema_selector.set(self.schema_name or 'public')
            self.schema_selector.bind("<<ComboboxSelected>>", self._on_schema_changed)

        # AI 聊天面板切换按钮
        ttk.Separator(self.toolbar, orient="vertical").pack(side="left", padx=10, fill="y")
        self.ai_toggle_btn = ttk.Button(self.toolbar, text="🧠 AI ▶", command=self._toggle_ai_panel)
        self.ai_toggle_btn.pack(side="right", padx=2)

        # 水平 PanedWindow：左侧编辑区+结果区，右侧 AI 聊天面板
        self.horizontal_paned = ttk.PanedWindow(self, orient="horizontal")
        self.horizontal_paned.pack(fill="both", expand=True, padx=5, pady=5)

        # 左侧：垂直分栏（编辑器+结果）
        left_frame = tk.Frame(self.horizontal_paned)
        self.horizontal_paned.add(left_frame, weight=2)

        self.main_paned = ttk.PanedWindow(left_frame, orient="vertical")
        self.main_paned.pack(fill="both", expand=True)
        
        self.editor_frame = tk.Frame(self.main_paned)
        self.main_paned.add(self.editor_frame, weight=2)
        
        self._init_find_bar(self.editor_frame)
        
        self.editor = scrolledtext.ScrolledText(self.editor_frame, undo=True, wrap="word", font=("Consolas", 12))
        self.editor.pack(fill="both", expand=True)
        
        self.editor.bind("<KeyRelease>", self._on_key_release)
        self.editor.bind("<FocusIn>", lambda e: self._hide_completions())
        self.editor.bind("<Button-1>", lambda e: self._hide_completions())
        self.editor.bind("<Tab>", self._on_tab_press)
        self.editor.bind("<Return>", self._on_enter_press)
        self.editor.bind("<Up>", lambda e: self._move_comp_selection(-1))
        self.editor.bind("<Down>", lambda e: self._move_comp_selection(1))
        self.editor.bind("<F5>", lambda e: self._run_query(needs_confirm=True))
        self.editor.bind("<Control-Return>", lambda e: self._run_query(needs_confirm=True))
        self.editor.bind("<Control-f>", lambda e: self._show_find_bar())
        self.editor.bind("<Control-s>", lambda e: self._save_script())
        self.editor.bind("<Control-space>", lambda e: self._show_completions(manual=True))
        self.editor.bind("<<Modified>>", self._on_editor_modified)
        self.editor.bind("<Alt-slash>", lambda e: self._show_completions(manual=True))
        self.editor.bind("<Escape>", self._on_esc_pressed)
        self.editor.bind("<<Paste>>", self._on_paste)
        
        self.editor_menu = tk.Menu(self, tearoff=0)
        self.editor_menu.add_command(label="▶ 执行选中", command=lambda: self._run_query(needs_confirm=False))
        self.editor_menu.add_separator()
        self.editor_menu.add_command(label="撤销", command=lambda: self.editor.event_generate("<<Undo>>"))
        self.editor_menu.add_separator()
        self.editor_menu.add_command(label="剪切", command=lambda: self.editor.event_generate("<<Cut>>"))
        self.editor_menu.add_command(label="复制", command=lambda: self.editor.event_generate("<<Copy>>"))
        self.editor_menu.add_command(label="粘贴", command=lambda: self.editor.event_generate("<<Paste>>"))
        self.editor_menu.add_command(label="全选", command=lambda: self.editor.event_generate("<<SelectAll>>"))
        self.editor.bind("<Button-3>", self._show_editor_context_menu)
        self.editor.focus_set()
        
        self.res_outer_frame = tk.Frame(self.main_paned)
        self.main_paned.add(self.res_outer_frame, weight=1)
        
        self.result_notebook = ttk.Notebook(self.res_outer_frame)
        self.result_notebook.pack(side="top", fill="both", expand=True)
        
        self.res_frame = tk.Frame(self.result_notebook)
        self.result_notebook.add(self.res_frame, text="查询结果")
        self.res_frame.grid_columnconfigure(0, weight=1)
        self.res_frame.grid_rowconfigure(0, weight=1)
        
        self.tree = ttk.Treeview(self.res_frame, show="headings", selectmode="extended")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Double-1>", self._on_tree_cell_double_click)
        self.tree.tag_configure("modified", background="#fff3cd", foreground="#856404")
        
        sy = ttk.Scrollbar(self.res_frame, orient="vertical", command=self.tree.yview)
        sx = ttk.Scrollbar(self.res_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")

        # --- 数据操作栏 (下沉到结果集下方) ---
        _c = get_theme_colors(self.settings.get('theme', '默认'))
        hb, bg, fg = _c["hb"], _c["bg"], _c["fg"]
        self.data_footer = tk.Frame(self.res_frame, padx=5, pady=2, bg=hb, borderwidth=1, relief="flat")
        self.data_footer.grid(row=2, column=0, columnspan=2, sticky="ew")
        
        # 左侧：数据编辑按钮
        self.save_data_btn = ttk.Button(self.data_footer, text="✓ 保存修改", command=self._save_data_changes, state="disabled")
        self.save_data_btn.pack(side="left", padx=2)
        
        self.cancel_data_btn = ttk.Button(self.data_footer, text="✗ 取消修改", command=self._cancel_data_changes, state="disabled")
        self.cancel_data_btn.pack(side="left", padx=2)

        ttk.Separator(self.data_footer, orient="vertical").pack(side="left", padx=10, fill="y")

        # 中间：分页控制
        self.page_prev_btn = ttk.Button(self.data_footer, text="◀ 上一页", width=8, command=self._prev_page, state="disabled")
        self.page_prev_btn.pack(side="left", padx=2)
        
        self.page_info_label = tk.Label(self.data_footer, text="第 1 / 1 页", font=("Microsoft YaHei", 9), bg=hb)
        self.page_info_label.pack(side="left", padx=5)
        
        self.page_next_btn = ttk.Button(self.data_footer, text="下一页 ▶", width=8, command=self._next_page, state="disabled")
        self.page_next_btn.pack(side="left", padx=2)
        
        tk.Label(self.data_footer, text="每页:", font=("Microsoft YaHei", 9), bg=hb).pack(side="left", padx=(10, 2))
        self.page_size_entry = tk.Entry(self.data_footer, textvariable=self.page_size, width=6, font=("Microsoft YaHei", 9))
        self.page_size_entry.pack(side="left", padx=2)
        self.page_size_entry.bind("<Return>", lambda e: self._render_current_page(reset_to_first=True))
        
        # 右侧：提示
        self.edit_tip_label = tk.Label(self.data_footer, text="双击单元格内联编辑", font=("Microsoft YaHei", 8), bg=hb, fg="#888")
        self.edit_tip_label.pack(side="right", padx=10)
        self.tree.configure(yscrollcommand=sy.set)
        sy.grid(row=0, column=1, sticky="ns")
        
        sx = ttk.Scrollbar(self.res_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(xscrollcommand=sx.set)
        sx.grid(row=1, column=0, sticky="ew")
        
        self.res_menu = tk.Menu(self, tearoff=0)
        self.res_menu.add_command(label="复制单元格", command=self._copy_cell)
        self.res_menu.add_command(label="复制整行 (仅数据)", command=self._copy_row)
        self.res_menu.add_command(label="复制选中 (含标题)", command=lambda: self._copy_selected_with_headers())
        self.res_menu.add_separator()
        self.res_menu.add_command(label="复制全部 (含标题)", command=lambda: self._copy_all_results(include_headers=True))
        self.res_menu.add_command(label="复制全部 (仅数据)", command=lambda: self._copy_all_results(include_headers=False))
        self.res_menu.add_separator()
        self.res_menu.add_command(label="复制为 INSERT 语句", command=self._copy_as_insert)
        self.res_menu.add_command(label="复制为 UPDATE 语句", command=self._copy_as_update)
        self.res_menu.add_separator()
        self.res_menu.add_command(label="导出为 CSV...", command=self._export_results_csv)
        self.res_menu.add_command(label="导出为 Excel...", command=self._export_results_excel)
        self.tree.bind("<Button-3>", self._show_header_context_menu)
        
        # 列头右键菜单（排序/筛选）
        self.header_menu = tk.Menu(self, tearoff=0)
        self.header_menu.add_command(label="按此列升序排序 ↑", command=lambda: self._sort_results(self._context_col, True))
        self.header_menu.add_command(label="按此列降序排序 ↓", command=lambda: self._sort_results(self._context_col, False))
        self.header_menu.add_separator()
        self.header_menu.add_command(label="筛选...", command=self._filter_results)
        self.header_menu.add_command(label="清除筛选", command=self._clear_filter)
        self._context_col = None  # 右键点击的列名
        
        self.msg_log_frame = tk.Frame(self.result_notebook)
        self.result_notebook.add(self.msg_log_frame, text="消息日志")
        
        self.msg_toolbar = tk.Frame(self.msg_log_frame, pady=2, bg=get_theme_colors(self.settings.get('theme', '默认'))["hb"])
        self.msg_toolbar.pack(side="top", fill="x")
        
        ttk.Button(self.msg_toolbar, text="📂 导出日志", command=self._export_log).pack(side="left", padx=5)
        ttk.Button(self.msg_toolbar, text="🗑️ 清空日志", command=self._clear_log).pack(side="left", padx=5)
        
        self.msg_area = scrolledtext.ScrolledText(self.msg_log_frame, state="disabled", font=("Microsoft YaHei", 9))
        self.msg_area.pack(fill="both", expand=True)

        # AI 聊天面板（右侧）
        from ui.ai.chat_panel import AIChatPanel
        self.chat_panel = AIChatPanel(
            self.horizontal_paned,
            conn_data=self.conn_data,
            db_ops=self.db_ops,
            db_name=self.db_name,
            schema_name=self.schema_name,
            conn_id=self.conn_data.get('id'),
            script_path=self.script_path
        )
        # 默认不添加到 PanedWindow (折叠状态)
        # self.horizontal_paned.add(self.chat_panel, weight=1)

        # AI 面板状态
        self._ai_panel_visible = False
        # 默认宽度设为屏幕宽度的 1/4 作为保底，展开时将遵循 2:1 权重分配
        self._ai_panel_width = self.winfo_screenwidth() // 4

    def _export_log(self):
        c = self.msg_area.get("1.0", tk.END).strip()
        if not c:
            messagebox.showinfo("提示", "日志为空")
            return
        p = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
        if p:
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(c)
                messagebox.showinfo("成功", f"导出至:\n{p}")
            except Exception as e:
                messagebox.showerror("失败", str(e))

    def _clear_log(self):
        if messagebox.askyesno("确认", "清空日志？"):
            self.msg_area.configure(state="normal")
            self.msg_area.delete("1.0", tk.END)
            self.msg_area.configure(state="disabled")

    def _show_editor_context_menu(self, event):
        try:
            has_sel = bool(self.editor.tag_ranges(tk.SEL))
            state = "normal" if has_sel else "disabled"
            # 批量设置与选中相关的菜单项状态
            self.editor_menu.entryconfig("▶ 执行选中", state=state)
            self.editor_menu.entryconfig("剪切", state=state)
            self.editor_menu.entryconfig("复制", state=state)
        except:
            pass
        self.editor_menu.post(event.x_root, event.y_root)

    def _toggle_ai_panel(self):
        """切换 AI 聊天面板的显示/隐藏"""
        if self._ai_panel_visible:
            # 折叠：记录宽度，移除面板
            try:
                self._ai_panel_width = self.chat_panel.winfo_width()
            except Exception:
                self._ai_panel_width = self.winfo_screenwidth() // 4
            self.horizontal_paned.forget(self.chat_panel)
            self._ai_panel_visible = False
            self.ai_toggle_btn.configure(text="🤖 AI ▶")
        else:
            # 展开：恢复面板
            if self._ai_panel_width > 10:
                self.chat_panel.configure(width=self._ai_panel_width)
            self.horizontal_paned.add(self.chat_panel, weight=1)
            self._ai_panel_visible = True
            self.ai_toggle_btn.configure(text="🤖 AI ◀")

    def _on_key_release(self, event):
        # 语法高亮防抖处理
        if hasattr(self, '_highlight_timer'):
            self.after_cancel(self._highlight_timer)
        self._highlight_timer = self.after(300, self._highlight_syntax)

        # 忽略导航键
        if event.keysym in ("Up", "Down", "Left", "Right", "Escape", "Return", "Tab"):
            return
        # 忽略 Ctrl 组合键
        if event.state & 0x4 and event.keysym.lower() not in ('v', 'x'):
            return

        idx = self.editor.index(tk.INSERT)
        line = self.editor.get(f"{idx} linestart", idx)
        tokens = re.split(r'[\s\n\(\),;]', line)
        curr = tokens[-1] if tokens else ""

        # Navicat 风格触发逻辑:
        # 1. 点号触发 (table. / alias.)
        if line.endswith('.'):
            self._show_completions(manual=False)
            return

        # 2. 输入 2+ 个字母时自动触发
        if len(curr) >= 2:
            self._show_completions(manual=False)
            return

        # 3. 输入 1 个字母时也触发（让用户能快速看到候选）
        if len(curr) == 1:
            self._show_completions(manual=False)
            return

        # 4. 当前 token 为空 → 隐藏补全
        self._hide_completions()

    def _show_completions(self, manual=False):
        idx = self.editor.index(tk.INSERT)
        # 取整个 SQL 上下文（从语句开头到光标），支持嵌套查询
        text_before_cursor = self._get_sql_context(idx)
        # 取完整 SQL 语句（到语句结尾），用于解析光标后方的别名定义
        full_statement = self._get_full_statement(idx)
        # 同步当前数据库到 completer
        self.completer.current_db = self._current_db
        sugs = self.completer.get_completions(text_before_cursor, full_statement)
        if not sugs:
            self._hide_completions()
            return

        # 单条建议且非手动触发时：完全匹配则隐藏弹窗
        if not manual and len(sugs) == 1:
            line = self.editor.get(f"{idx} linestart", idx)
            tokens = re.split(r'[\s\n\(\),;]', line)
            last_token = tokens[-1] if tokens else ""
            # 获取建议的最终插入文本
            sug_insert = sugs[0].insert_text if isinstance(sugs[0], CompletionItem) else (
                sugs[0][0] if isinstance(sugs[0], tuple) else sugs[0]
            )
            if last_token.upper() == sug_insert.upper():
                self._hide_completions()
                return

        # 创建补全弹窗
        if not self.comp_window or not self.comp_window.winfo_exists():
            self.comp_window = tk.Toplevel(self)
            self.comp_window.overrideredirect(True)
            self.comp_window.attributes("-topmost", True)
            self.comp_listbox = tk.Listbox(
                self.comp_window, height=10, font=("Microsoft YaHei", 9),
                borderwidth=1, relief="solid", exportselection=False
            )
            self.comp_listbox.pack(fill="both", expand=True)
            self.comp_listbox.bind("<Double-1>", lambda e: self._apply_completion())

        self.comp_listbox.delete(0, tk.END)

        # 存储原始补全项
        self._completion_items = []
        display_items = []
        max_display_len = 0

        for s in sugs:
            if isinstance(s, CompletionItem):
                self._completion_items.append(s)
                # 格式化显示文本: name + (detail)
                if s.type == CompletionItem.TYPE_COLUMN and s.detail:
                    display_text = f"{s.display}  ({s.detail})"
                elif s.type == CompletionItem.TYPE_TABLE and s.detail:
                    display_text = f"{s.display}  -- {s.detail}"
                elif s.type == CompletionItem.TYPE_FUNCTION:
                    display_text = f"{s.display}  {s.detail}"
                elif s.type == CompletionItem.TYPE_ALIAS and s.detail:
                    display_text = f"{s.display}  {s.detail}"
                else:
                    display_text = s.display
                display_items.append((s, display_text))
                self.comp_listbox.insert(tk.END, display_text)
                max_display_len = max(max_display_len, len(display_text))
            elif isinstance(s, tuple):
                # 向后兼容: (name, comment)
                name, comment = s
                self._completion_items.append(s)
                display_text = f"{name}  -- {comment}" if comment else name
                display_items.append((name, display_text))
                self.comp_listbox.insert(tk.END, display_text)
                max_display_len = max(max_display_len, len(display_text))
            else:
                # 向后兼容: plain string
                self._completion_items.append(s)
                display_items.append((s, s))
                self.comp_listbox.insert(tk.END, s)
                max_display_len = max(max_display_len, len(s))

        if display_items:
            self.comp_listbox.selection_set(0)

        # 自适应弹窗宽度和高度
        popup_w = max(260, min(500, max_display_len * 8 + 40))
        popup_h = min(250, len(sugs) * 20 + 10)
        bbox = self.editor.bbox(tk.INSERT)
        if bbox:
            x, y, w, h = bbox
            rx, ry = self.editor.winfo_rootx(), self.editor.winfo_rooty()
            self.comp_window.geometry(f"{popup_w}x{popup_h}+{rx + x}+{ry + y + h + 2}")
            self.comp_window.deiconify()

    def _get_sql_context(self, cursor_idx):
        """获取从当前 SQL 语句开头到光标的完整文本，支持嵌套查询上下文识别"""
        # 从光标向前找语句起始（以分号或编辑器开头为界）
        row, col = map(int, cursor_idx.split('.'))
        # 向前搜索分号，确定语句起点
        start_idx = "1.0"
        search_text = self.editor.get("1.0", cursor_idx)
        last_semicolon = search_text.rfind(';')
        if last_semicolon >= 0:
            # 计算分号后的位置
            before_semi = search_text[:last_semicolon + 1]
            lines = before_semi.count('\n')
            if lines == 0:
                start_col = last_semicolon + 1
                start_idx = f"1.{start_col}"
            else:
                last_nl = before_semi.rfind('\n')
                start_col = len(before_semi) - last_nl - 1
                start_idx = f"{lines + 1}.{start_col}"
        return self.editor.get(start_idx, cursor_idx)

    def _get_full_statement(self, cursor_idx):
        """获取光标所在完整 SQL 语句（从语句开头到语句结尾），用于解析光标后方的别名定义"""
        # 先找到语句起始（复用 _get_sql_context 的逻辑）
        start_idx = "1.0"
        search_text = self.editor.get("1.0", cursor_idx)
        last_semicolon = search_text.rfind(';')
        if last_semicolon >= 0:
            before_semi = search_text[:last_semicolon + 1]
            lines = before_semi.count('\n')
            if lines == 0:
                start_col = last_semicolon + 1
                start_idx = f"1.{start_col}"
            else:
                last_nl = before_semi.rfind('\n')
                start_col = len(before_semi) - last_nl - 1
                start_idx = f"{lines + 1}.{start_col}"

        # 从光标向后找语句结尾（下一个分号或编辑器末尾）
        end_idx = self.editor.index(tk.END)
        after_cursor = self.editor.get(cursor_idx, tk.END)
        next_semicolon = after_cursor.find(';')
        if next_semicolon >= 0:
            # 计算分号位置
            before_end = after_cursor[:next_semicolon + 1]
            lines_after = before_end.count('\n')
            row, col = map(int, cursor_idx.split('.'))
            if lines_after == 0:
                end_idx = f"{row}.{col + next_semicolon + 1}"
            else:
                last_nl = before_end.rfind('\n')
                end_col = len(before_end) - last_nl - 1
                end_idx = f"{row + lines_after}.{end_col}"

        return self.editor.get(start_idx, end_idx)

    def _hide_completions(self):
        if self.comp_window and self.comp_window.winfo_exists():
            self.comp_window.withdraw()

    def _move_comp_selection(self, delta):
        if self.comp_window and self.comp_window.winfo_viewable():
            curr = self.comp_listbox.curselection()
            if curr:
                n = max(0, min(self.comp_listbox.size()-1, curr[0] + delta))
                self.comp_listbox.selection_clear(0, tk.END)
                self.comp_listbox.selection_set(n)
                self.comp_listbox.see(n)
            return "break"

    def _on_tab_press(self, event):
        if self.comp_window and self.comp_window.winfo_viewable():
            self._apply_completion()
            return "break"

    def _on_esc_pressed(self, event=None):
        self._hide_completions()
        if hasattr(self, 'find_frame') and self.find_frame.winfo_viewable():
            self._hide_find_bar()
        return "break"

    def _on_enter_press(self, event):
        if self.comp_window and self.comp_window.winfo_viewable():
            self._apply_completion()
            return "break"
        if self.editor.tag_ranges(tk.SEL):
            self._run_query(needs_confirm=True)
            return "break"

    def _run_query(self, needs_confirm=False):
        # 增加执行确认 (仅当 needs_confirm 为 True 时触发，如快捷键)
        if needs_confirm:
            if not messagebox.askokcancel("确认执行", "确定要执行当前的 SQL 脚本吗？"):
                return

        try:
            sql = self.editor.get(tk.SEL_FIRST, tk.SEL_LAST).strip()
        except:
            sql = self.editor.get("1.0", tk.END).strip()
        if not sql:
            return
            
        self.is_running = True
        
        self.exec_btn.config(text="⌛ 执行中...", state="disabled")
        self.stop_btn.config(state="normal")
        
        # 记录执行日志
        self._log_msg(f"开始执行 SQL (库: {self.db_name}, 模式: {self.schema_name or '默认'}):")
        self._log_msg(f"  {sql}", "INFO")
        
        self.current_query_table = self._detect_table_name(sql)
        self.current_query_pks = []
        self.pending_updates.clear()
        self.original_values.clear()
        self.save_data_btn.config(state="disabled")
        self.cancel_data_btn.config(state="disabled")
        
        progress = ProgressDialog(self, message="正在准备执行...")
        def run():
            try:
                if self.current_query_table:
                    self.current_query_pks = self.db_ops.get_primary_keys(
                        self.conn_data, self.current_query_table,
                        database=self.db_name, schema=self.schema_name
                    )
                
                if self.cancel_event: self.cancel_event.clear()
                
                # 定义进度回调
                def on_progress(pct, msg):
                    self.after(0, lambda: progress.update_progress(pct, msg))

                # 传递 schema 参数，execute_sql 会在同一连接内先 SET search_path
                schema = self.schema_name if self.conn_data.get('db_type') == 'PostgreSQL' else None
                cols, rows, aff, isq = self.db_ops.execute_sql(
                    self.conn_data, sql, database=self.db_name,
                    cancel_event=self.cancel_event, schema=schema,
                    progress_callback=on_progress
                )
                self.after(0, lambda: self._on_execute_done(progress, cols, rows, aff, isq, None, sql))
            except Exception as e:
                self.after(0, lambda e=e: self._on_execute_done(progress, [], [], 0, False, str(e), sql))
        threading.Thread(target=run, daemon=True).start()

    def _stop_query(self):
        """中止当前正在执行的 SQL 查询"""
        if self.cancel_event:
            self._log_msg("收到停止指令，正在尝试中止执行...", "WARN")
            self.cancel_event.set()
        
        self.stop_btn.config(state="disabled")
        self.is_running = False

    def _on_execute_done(self, progress, cols, rows, affected, is_query, error, original_sql):
        progress.close()
        self.is_running = False
             
        self.exec_btn.config(text="▶ 执行 (F5)", state="normal")
        self.stop_btn.config(state="disabled")
        
        if error:
            self._log_msg(f"执行失败: {error}", "ERROR")
            self.result_notebook.select(1)
            return
        if is_query:
            if self.current_query_table and self.current_query_pks:
                self.edit_tip_label.config(text=f"表: {self.current_query_table} (双击编辑)", fg="blue")
            else:
                self.edit_tip_label.config(text="结果只读", fg="gray")
            self._display_results(cols, rows)
            self.result_notebook.select(0)
        else:
            self._log_msg(f"影响 {affected} 行")
            self.result_notebook.select(1)

    def _init_find_bar(self, parent):
        _c = get_theme_colors(self.settings.get('theme', '默认'))
        hb, bg, fg = _c["hb"], _c["bg"], _c["fg"]
        
        self.find_frame = tk.Frame(parent, bg=hb, pady=2)
        
        tk.Label(self.find_frame, text="🔍 查找:", bg=hb).pack(side="left", padx=5)
        self.find_entry = tk.Entry(self.find_frame, width=30)
        self.find_entry.pack(side="left", padx=5)
        self.find_entry.bind("<Return>", lambda e: self._find_next())
        self.find_entry.bind("<KeyRelease>", lambda e: self._do_find())
        
        self.find_stats = tk.Label(self.find_frame, text="0/0", bg=hb, width=10)
        self.find_stats.pack(side="left", padx=5)
        
        ttk.Button(self.find_frame, text="↑", width=3, command=self._find_prev).pack(side="left", padx=2)
        ttk.Button(self.find_frame, text="↓", width=3, command=self._find_next).pack(side="left", padx=2)
        
        tk.Button(self.find_frame, text="✕", command=self._hide_find_bar, relief="flat", bg=hb).pack(side="right", padx=5)

    def _show_find_bar(self):
        if not self.find_frame.winfo_viewable():
            self.find_frame.pack(side="top", fill="x", before=self.editor)
        self.find_entry.focus_set()
        self.find_entry.select_range(0, tk.END)
        self._do_find()

    def _hide_find_bar(self):
        self.find_frame.pack_forget()
        self.editor.tag_remove("find_match", "1.0", tk.END)
        self.editor.focus_set()

    def _do_find(self):
        self.editor.tag_remove("find_match", "1.0", tk.END)
        query = self.find_entry.get()
        if not query:
            self.find_stats.config(text="0/0")
            return
        
        count = 0
        start = "1.0"
        while True:
            start = self.editor.search(query, start, stopindex=tk.END, nocase=True)
            if not start:
                break
            end = f"{start}+{len(query)}c"
            self.editor.tag_add("find_match", start, end)
            start = end
            count += 1
        
        self.editor.tag_config("find_match", background="yellow", foreground="black")
        self.find_stats.config(text=f"共 {count} 个")

    def _find_next(self):
        query = self.find_entry.get()
        if not query: return
        curr = self.editor.index(tk.INSERT)
        start = self.editor.search(query, f"{curr}+1c", stopindex=tk.END, nocase=True)
        if not start:
            start = self.editor.search(query, "1.0", stopindex=tk.END, nocase=True)
        if start:
            self.editor.mark_set(tk.INSERT, start)
            self.editor.see(start)

    def _find_prev(self):
        query = self.find_entry.get()
        if not query: return
        curr = self.editor.index(tk.INSERT)
        start = self.editor.search(query, curr, stopindex="1.0", nocase=True, backwards=True)
        if not start:
            start = self.editor.search(query, tk.END, stopindex="1.0", nocase=True, backwards=True)
        if start:
            self.editor.mark_set(tk.INSERT, start)
            self.editor.see(start)

    def _display_results(self, cols, rows):
        """初始化全量结果并显示第一页"""
        self.all_cols = cols
        self.all_rows = rows
        # 新查询结果抵达，重置排序/筛选状态
        self.sort_column = None
        self.sort_ascending = True
        self.filter_column = None
        self.filter_value = ""
        self.original_all_rows = []
        self._render_current_page(reset_to_first=True)

    def _render_current_page(self, reset_to_first=False):
        """渲染当前页的数据到 Treeview"""
        if reset_to_first:
            self.current_page = 1
            
        try:
            ps = int(self.page_size.get())
            if ps <= 0: ps = 5000
        except:
            ps = 5000
            self.page_size.set(ps)
            
        total_rows = len(self.all_rows)
        self.total_pages = max(1, (total_rows + ps - 1) // ps)
        
        if self.current_page > self.total_pages:
            self.current_page = self.total_pages
            
        # 更新 UI 状态
        self.page_info_label.config(text=f"第 {self.current_page} / {self.total_pages} 页 (共 {total_rows} 条数据)")
        self.page_prev_btn.config(state="normal" if self.current_page > 1 else "disabled")
        self.page_next_btn.config(state="normal" if self.current_page < self.total_pages else "disabled")

        # 清理旧数据和状态
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 重新配置列
        self.tree["columns"] = self.all_cols
        for col in self.all_cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, stretch=False)
        self._render_header_indicators()
            
        # 计算切片
        start_idx = (self.current_page - 1) * ps
        end_idx = min(start_idx + ps, total_rows)
        
        if total_rows == 0:
            self._log_msg("查询成功，但结果集为空。")
            return

        # 预分配 all_rows_data，避免逐条 append 动态扩容
        page_size = end_idx - start_idx
        self.all_rows_data = [None] * page_size

        # 内部采用分批渲染，防止切换分页时瞬间卡死
        # 放大 batch_size 减少 Tcl/Tk 跨语言调用次数；after_idle 替代 after(5)
        batch_size = 2000
        
        # 全局只需判断一次 row 类型，后续所有 batch 共用
        is_dict_rows = bool(self.all_rows) and isinstance(self.all_rows[0], dict)
        # 局部变量引用加速循环内属性访问
        all_rows = self.all_rows
        all_cols = self.all_cols
        clean_val = self._clean_value
        tree = self.tree
        base_offset = start_idx
        
        def load_inner_batch(curr_start):
            if not self.winfo_exists():
                return
            curr_end = min(curr_start + batch_size, end_idx)
            
            if is_dict_rows:
                for i in range(curr_start, curr_end):
                    row = all_rows[i]
                    cv = [clean_val(row.get(c, "")) for c in all_cols]
                    self.all_rows_data[i - base_offset] = cv
                    tag = ("odd",) if i % 2 else ()
                    tree.insert("", tk.END, values=cv, tags=tag)
            else:
                for i in range(curr_start, curr_end):
                    row = all_rows[i]
                    cv = [clean_val(v) for v in row]
                    self.all_rows_data[i - base_offset] = cv
                    tag = ("odd",) if i % 2 else ()
                    tree.insert("", tk.END, values=cv, tags=tag)
                
            if curr_end < end_idx:
                self.after_idle(lambda: load_inner_batch(curr_end))
            else:
                self._log_msg(f"正在显示第 {start_idx+1} 到 {end_idx} 条数据 (共 {total_rows} 行)")

        load_inner_batch(start_idx)

    def _prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._render_current_page()

    def _next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self._render_current_page()

    def _show_header_context_menu(self, event):
        """列头右键菜单：排序/筛选"""
        region = self.tree.identify_region(event.x, event.y)
        if region != "heading":
            self._show_result_context_menu(event)
            return
        col_id = self.tree.identify_column(event.x)  # "#1", "#2", ...
        if not col_id:
            return
        try:
            idx = int(col_id.replace("#", "")) - 1
            cols = self.tree["columns"]
            if 0 <= idx < len(cols):
                self._context_col = cols[idx]
                # 更新菜单状态
                filter_col_name = self.filter_column
                self.header_menu.entryconfig(4, label=f"清除筛选 ({filter_col_name})" if filter_col_name else "清除筛选")
                self.header_menu.post(event.x_root, event.y_root)
        except Exception as e:
            self._log_msg(f"列头菜单出错: {e}", "ERROR")

    def _sort_results(self, col_name, ascending=True):
        """对结果集按指定列排序"""
        if not self.all_rows or not self.all_cols:
            return
        try:
            idx = self.all_cols.index(col_name)
        except ValueError:
            self._log_msg(f"找不到列: {col_name}", "ERROR")
            return
        
        if self.sort_column == col_name and self.sort_ascending == ascending:
            return  # 已经是相同排序
        
        # 确保列值可比较
        rows = self.all_rows
        if rows and isinstance(rows[0], dict):
            rows.sort(key=lambda r: (str(r.get(col_name, "")) or ""), reverse=not ascending)
        else:
            rows.sort(key=lambda r: (str(r[idx]) if idx < len(r) else ""), reverse=not ascending)
        
        self.sort_column = col_name
        self.sort_ascending = ascending
        self._render_header_indicators()
        self._log_msg(f"已按「{col_name}」{'升序' if ascending else '降序'}排序")
        self._render_current_page()

    def _filter_results(self):
        """筛选结果集"""
        if not self.all_rows or not self.all_cols:
            return
        col_name = self._context_col
        if not col_name:
            return
        from tkinter.simpledialog import askstring
        val = askstring("筛选", f"输入筛选值（包含即匹配）:\n列: {col_name}", parent=self)
        if val is None:
            return
        
        # 首次筛选时备份原始数据
        if not self.original_all_rows:
            self.original_all_rows = self.all_rows[:]
        
        try:
            idx = self.all_cols.index(col_name)
        except ValueError:
            return
        
        val_lower = val.lower().strip()
        rows = self.original_all_rows if self.original_all_rows else self.all_rows
        if rows and isinstance(rows[0], dict):
            filtered = [r for r in rows if val_lower in (str(r.get(col_name, "")) or "").lower()]
        else:
            filtered = [r for r in rows if val_lower in (str(r[idx]) if idx < len(r) else "").lower()]
        
        self.all_rows = filtered
        self.filter_column = col_name
        self.filter_value = val
        self._update_pagination()
        self._log_msg(f"已筛选「{col_name}」包含「{val}」，匹配 {len(filtered)} 行")
        self._render_current_page()

    def _clear_filter(self):
        """清除筛选，恢复原始数据"""
        if not self.original_all_rows:
            return
        self.all_rows = self.original_all_rows[:]
        self.original_all_rows = []
        self.filter_column = None
        self.filter_value = ""
        self._update_pagination()
        self._log_msg("已清除筛选")
        self._render_current_page()

    def _render_header_indicators(self):
        """在列头显示排序指示符"""
        cols = self.tree["columns"]
        for col in cols:
            text = col
            if self.sort_column == col:
                text = f"{'↑' if self.sort_ascending else '↓'} {col}"
            self.tree.heading(col, text=text)

    def _update_pagination(self):
        """更新分页信息"""
        total = len(self.all_rows)
        page_size = self.page_size.get()
        self.total_pages = max(1, (total + page_size - 1) // page_size)
        self.current_page = min(self.current_page, self.total_pages)

    def _show_result_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.res_menu.post(event.x_root, event.y_root)

    def _format_sql(self):
        """美化 SQL 代码"""
        try:
            import sqlparse
            # 优先美化选中的内容
            try:
                sql = self.editor.get(tk.SEL_FIRST, tk.SEL_LAST).strip()
                is_selection = True
            except:
                sql = self.editor.get("1.0", tk.END).strip()
                is_selection = False
                
            if not sql: return
            
            formatted = sqlparse.format(sql, reindent=True, keyword_case='upper')
            
            if is_selection:
                self.editor.delete(tk.SEL_FIRST, tk.SEL_LAST)
                self.editor.insert(tk.INSERT, formatted)
            else:
                self.editor.delete("1.0", tk.END)
                self.editor.insert("1.0", formatted)
                
            self._highlight_syntax()
            self._log_msg("SQL 已美化")
        except ImportError:
            messagebox.showwarning("提示", "请安装 sqlparse 库以使用美化功能:\npip install sqlparse")
        except Exception as e:
            self._log_msg(f"美化失败: {e}", "ERROR")

    def _copy_cell(self):
        """复制选中的单元格内容"""
        sel = self.tree.selection()
        if not sel:
            return
        # 获取鼠标位置对应的列
        x = self.tree.winfo_pointerx() - self.tree.winfo_rootx()
        cid = self.tree.identify_column(x)
        try:
            # 转换为 0-based 索引
            idx = int(cid.replace("#", "")) - 1
            vals = self.tree.item(sel[0])['values']
            if 0 <= idx < len(vals):
                val = vals[idx]
                self.clipboard_clear()
                self.clipboard_append(str(val))
                self._log_msg(f"已复制单元格: {val}")
        except Exception as e:
            self._log_msg(f"复制失败: {e}", "ERROR")

    def _copy_row(self):
        """复制选中的整行内容（制表符分隔）"""
        sel = self.tree.selection()
        if not sel:
            return
        
        rows_text = []
        for item in sel:
            vals = self.tree.item(item)['values']
            rows_text.append("\t".join([str(v) for v in vals]))
            
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows_text))
        self._log_msg(f"已复制 {len(sel)} 行数据")

    def _copy_selected_with_headers(self):
        """复制选中的行并包含标题"""
        sel = self.tree.selection()
        if not sel: return
        
        cols = self.tree["columns"]
        lines = ["\t".join(cols)]
        
        for item in sel:
            vals = self.tree.item(item)['values']
            lines.append("\t".join([str(v) for v in vals]))
            
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))
        self._log_msg(f"已复制选中数据 (含标题, 共 {len(sel)} 行)")

    def _copy_all_results(self, include_headers=True):
        """复制所有查询结果到剪贴板"""
        cols = self.tree["columns"]
        lines = []
        if include_headers:
            lines.append("\t".join(cols))
        
        # 获取所有项
        all_items = self.tree.get_children()
        for item in all_items:
            vals = self.tree.item(item)['values']
            lines.append("\t".join([str(v) for v in vals]))
            
        if not lines: return
            
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))
        self._log_msg(f"已复制全部结果 ({'含标题' if include_headers else '仅数据'}, 共 {len(all_items)} 行)")

    def _copy_as_insert(self):
        """将选中行复制为 INSERT 语句"""
        if not self.current_query_table:
            messagebox.showwarning("提示", "未检测到单表上下文，无法生成 INSERT 语句")
            return
            
        sel = self.tree.selection()
        if not sel: return
        
        db_type = self.conn_data.get('db_type', 'MySQL')
        quote = "`" if db_type == "MySQL" else "\""
        table_name = f"{quote}{self.current_query_table}{quote}"
        if db_type == "PostgreSQL" and self.schema_name:
            table_name = f"\"{self.schema_name}\".\"{self.current_query_table}\""
            
        cols = self.tree["columns"]
        col_str = ", ".join([f"{quote}{c}{quote}" for c in cols])
        
        sqls = []
        for item in sel:
            vals = self.tree.item(item)['values']
            val_strs = []
            for v in vals:
                if v is None or v == "":
                    val_strs.append("NULL")
                elif isinstance(v, (int, float)):
                    val_strs.append(str(v))
                else:
                    # 简单转义单引号
                    escaped = str(v).replace("'", "''")
                    val_strs.append(f"'{escaped}'")
            
            sqls.append(f"INSERT INTO {table_name} ({col_str}) VALUES ({', '.join(val_strs)});")
            
        self.clipboard_clear()
        self.clipboard_append("\n".join(sqls))
        self._log_msg(f"已复制 {len(sqls)} 条 INSERT 语句")

    def _copy_as_update(self):
        """将选中行复制为 UPDATE 语句 (基于检测到的主键)"""
        if not self.current_query_table or not self.current_query_pks:
            messagebox.showwarning("提示", "未检测到表名或主键，无法生成精确的 UPDATE 语句")
            return
            
        sel = self.tree.selection()
        if not sel: return
        
        db_type = self.conn_data.get('db_type', 'MySQL')
        quote = "`" if db_type == "MySQL" else "\""
        table_name = f"{quote}{self.current_query_table}{quote}"
        if db_type == "PostgreSQL" and self.schema_name:
            table_name = f"\"{self.schema_name}\".\"{self.current_query_table}\""
            
        cols = self.tree["columns"]
        pk_indices = [cols.index(pk) for pk in self.current_query_pks if pk in cols]
        
        if not pk_indices:
            messagebox.showwarning("提示", "主键列不在查询结果中，无法定位行")
            return
            
        sqls = []
        for item in sel:
            vals = self.tree.item(item)['values']
            
            # SET 部分 (排除主键)
            set_clauses = []
            for i, col in enumerate(cols):
                if i in pk_indices: continue
                v = vals[i]
                v_str = "NULL" if (v is None or v == "") else (str(v) if isinstance(v, (int, float)) else "'{}'".format(str(v).replace("'", "''")))
                set_clauses.append(f"{quote}{col}{quote} = {v_str}")
                
            # WHERE 部分 (主键定位)
            where_clauses = []
            for idx in pk_indices:
                col = cols[idx]
                v = vals[idx]
                v_str = "NULL" if (v is None or v == "") else (str(v) if isinstance(v, (int, float)) else "'{}'".format(str(v).replace("'", "''")))
                where_clauses.append(f"{quote}{col}{quote} = {v_str}")
                
            sqls.append(f"UPDATE {table_name} SET {', '.join(set_clauses)} WHERE {' AND '.join(where_clauses)};")
            
        self.clipboard_clear()
        self.clipboard_append("\n".join(sqls))
        self._log_msg(f"已复制 {len(sqls)} 条 UPDATE 语句")

    def _clean_value(self, val):
        if val is None:
            return ""
        if isinstance(val, bytes):
            if val == b'\x00':
                return 0
            if val == b'\x01':
                return 1
            try:
                val = val.decode('utf-8')
            except:
                val = str(val)
        s = str(val)
        if s == "b'\x00'":
            return 0
        if s == "b'\x01'":
            return 1
        return _CLEAN_RE.sub('', s)

    def _export_results_csv(self):
        p = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
        if not p:
            return
        if not self.all_rows:
            messagebox.showwarning("提示", "没有数据可导出")
            return

        total = len(self.all_rows)
        is_dict = isinstance(self.all_rows[0], dict)
        cols = self.all_cols
        clean = self._clean_value

        progress = ProgressDialog(self, message=f"正在导出 {total} 行数据为 CSV...")
        progress.update_progress(0)

        def run():
            try:
                import csv
                with open(p, "w", encoding="utf-8-sig", newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow([clean(c) for c in cols])

                    update_step = max(100, min(5000, total // 200)) if total > 200 else 1

                    for i, row in enumerate(self.all_rows):
                        if is_dict:
                            writer.writerow([clean(row.get(c, "")) for c in cols])
                        else:
                            writer.writerow([clean(v) for v in row])
                        if i % update_step == 0:
                            pct = int((i + 1) / total * 100)
                            self.after(0, lambda v=pct: progress.update_progress(v))

                self.after(0, lambda: self._on_export_done(progress, True, p, "CSV"))
            except Exception as e:
                self.after(0, lambda e=e: self._on_export_done(progress, False, None, "CSV", e))

        threading.Thread(target=run, daemon=True).start()

    def _export_results_excel(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not p:
            return
        if not self.all_rows:
            messagebox.showwarning("提示", "没有数据可导出")
            return

        total = len(self.all_rows)
        is_dict = isinstance(self.all_rows[0], dict)
        cols = self.all_cols
        clean = self._clean_value
        use_write_only = total > 5000

        progress = ProgressDialog(self, message=f"正在导出 {total} 行数据为 Excel...")
        progress.update_progress(0)

        def run():
            try:
                from openpyxl import Workbook

                if use_write_only:
                    wb = Workbook(write_only=True)
                    ws = wb.create_sheet(title="Results")
                    ws.append([clean(c) for c in cols])
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Results"
                    ws.append([clean(c) for c in cols])
                    from openpyxl.styles import Font
                    for cell in ws[1]:
                        cell.font = Font(bold=True)

                update_step = max(100, min(5000, total // 200)) if total > 200 else 1

                if is_dict:
                    for i, row in enumerate(self.all_rows):
                        ws.append([clean(row.get(c, "")) for c in cols])
                        if i % update_step == 0:
                            pct = int((i + 1) / total * 100)
                            self.after(0, lambda v=pct: progress.update_progress(v))
                else:
                    for i, row in enumerate(self.all_rows):
                        ws.append([clean(v) for v in row])
                        if i % update_step == 0:
                            pct = int((i + 1) / total * 100)
                            self.after(0, lambda v=pct: progress.update_progress(v))

                wb.save(p)
                self.after(0, lambda: self._on_export_done(progress, True, p, "Excel"))
            except Exception as e:
                self.after(0, lambda e=e: self._on_export_done(progress, False, None, "Excel", e))

        threading.Thread(target=run, daemon=True).start()

    def _on_export_done(self, progress, success, path, fmt, error=None):
        """CSV/Excel 导出后台线程执行完毕后的 UI 回调"""
        progress.close()
        if success:
            messagebox.showinfo("成功", f"{fmt} 导出完成！\n{path}")
        else:
            messagebox.showerror("导出失败", str(error))

    def _save_script(self):
        """保存脚本逻辑：区分直接保存与另存为"""
        # 取消待执行的防抖自动保存，避免冗余
        if self._autosave_debounce_id:
            self.after_cancel(self._autosave_debounce_id)
            self._autosave_debounce_id = None
        content = self.editor.get("1.0", tk.END)
        
        # 检查当前是否为已有正式脚本（路径中不包含 autosave_）
        filename = os.path.basename(self.script_path)
        is_formal_script = "autosave_" not in filename
        
        if is_formal_script:
            # 直接保存到当前文件
            try:
                with open(self.script_path, "w", encoding="utf-8") as f:
                    f.write(content)
                self._log_msg(f"脚本已直接保存: {filename}")
                # 触发侧边栏刷新（如果需要）
                if self.on_save_callback:
                    self.on_save_callback()
            except Exception as e:
                messagebox.showerror("保存失败", str(e))
        else:
            # 新建查询，弹出命名对话框
            t = simpledialog.askstring("保存脚本", "请输入脚本名称:", parent=self)
            if t:
                # 记录旧日志内容以便迁移
                old_log = self.msg_area.get("1.0", tk.END)
                
                new_path = self.scratch_manager.save_script(self.conn_data['name'], self.db_name, t, content)
                if new_path:
                    old_path = self.script_path
                    # 删除旧的自动保存文件
                    if os.path.exists(old_path):
                        try: os.remove(old_path)
                        except: pass
                        
                    self.script_path = new_path
                    # 将当前日志保存到新路径
                    try:
                        self.storage.save_workbench_log(self.script_path, old_log)
                        # 同时迁移 AI 聊天历史
                        self.storage.migrate_chat_history(old_path, self.script_path)
                    except: pass

                    # 更新 AI 面板关联的路径
                    if hasattr(self, 'chat_panel'):
                        self.chat_panel.script_path = self.script_path
                    
                    # 更新标签页标题 (Notebook.tab)
                    try:
                        # self.master 是 Notebook
                        self.master.tab(self, text=f"Q: {t}")
                    except:
                        pass
                    
                    messagebox.showinfo("成功", f"脚本已保存为: {t}.sql")
                    self._log_msg(f"脚本已另存为正式文件: {t}.sql")
                    if self.on_save_callback:
                        self.on_save_callback()


    def _export_log(self):
        """导出日志"""
        content = self.msg_area.get("1.0", tk.END).strip()
        if not content:
            messagebox.showinfo("提示", "日志为空")
            return
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
        if path:
            try:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(content)
                messagebox.showinfo("成功", f"导出至:\n{path}")
            except Exception as e:
                messagebox.showerror("失败", str(e))

    def _clear_log(self):
        """清空日志并同步删除持久化数据"""
        if messagebox.askyesno("确认", "清空日志？"):
            self.msg_area.configure(state="normal")
            self.msg_area.delete("1.0", tk.END)
            self.msg_area.configure(state="disabled")
            # 持久化清空 (按路径)
            try:
                self.storage.delete_workbench_log(self.script_path)
            except: pass

    def _log_msg(self, message, level="INFO"):
        """向消息日志区域添加消息，并实时持久化 (按路径绑定)"""
        if not hasattr(self, 'msg_area') or not self.msg_area: return
        self.msg_area.configure(state="normal")
        time_str = datetime.now().strftime("%H:%M:%S")
        tag = level.upper()
        self.msg_area.insert(tk.END, f"[{time_str}] {tag}: {message}\n", tag)
        self.msg_area.see(tk.END)
        
        # 实时保存完整日志流
        full_content = self.msg_area.get("1.0", tk.END)
        self.msg_area.configure(state="disabled")
        
        # 异步保存 (按路径)
        def save():
            try:
                self.storage.save_workbench_log(self.script_path, full_content)
            except: pass
        threading.Thread(target=save, daemon=True).start()

        self.msg_area.tag_configure("ERROR", foreground="red")
        self.msg_area.tag_configure("WARN", foreground="orange")
        self.msg_area.tag_configure("INFO", foreground="black")

    def _detect_table_name(self, sql):
        """简单正则提取 SQL 中的表名，用于启用数据编辑"""
        match = re.search(r'FROM\s+[`"\[]?([\w\.]+)[`"\]]?', sql, re.IGNORECASE)
        if match:
            table = match.group(1)
            return table.split('.')[-1]
        return None

    def _on_tree_cell_double_click(self, event):
        """双击单元格进入内联编辑状态"""
        # 必须有表名和主键才能编辑
        if not self.current_query_table or not self.current_query_pks:
            return
            
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell": return
        
        column = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        if not item_id or not column: return
        
        # 获取列索引和名称
        try:
            col_idx = int(column.replace("#", "")) - 1
            cols = self.tree["columns"]
            if col_idx < 0 or col_idx >= len(cols): return
            col_name = cols[col_idx]
        except: return
        
        # 获取当前单元格的值
        item_data = self.tree.item(item_id)
        curr_vals = item_data.get("values", [])
        if not curr_vals or col_idx >= len(curr_vals): return
        curr_val = curr_vals[col_idx]
        
        # 计算位置并弹出 Entry
        bbox = self.tree.bbox(item_id, column)
        if not bbox: return
        x, y, w, h = bbox
        
        edit_entry = tk.Entry(self.tree, font=("Microsoft YaHei", 9), bd=1, relief="solid")
        edit_entry.insert(0, str(curr_val))
        edit_entry.select_range(0, tk.END)
        edit_entry.place(x=x, y=y, width=w, height=h)
        edit_entry.focus_set()
        
        # 状态标志，防止重复触发
        self._is_saving_edit = False

        def commit_change(event=None):
            if self._is_saving_edit or not edit_entry.winfo_exists(): return
            self._is_saving_edit = True
            
            new_val = edit_entry.get()
            # 比较字符串值，并去除两端空格
            if str(new_val).strip() != str(curr_val).strip():
                # 记录变更数据
                if item_id not in self.pending_updates:
                    self.pending_updates[item_id] = {}
                    self.original_values[item_id] = list(curr_vals)
                
                self.pending_updates[item_id][col_name] = new_val
                
                # 实时更新 Treeview 显示
                new_row_vals = list(self.tree.item(item_id, "values"))
                new_row_vals[col_idx] = new_val
                self.tree.item(item_id, values=new_row_vals, tags=("modified",))
                
                # 强制激活保存和取消按钮
                self.save_data_btn.config(state="normal")
                self.cancel_data_btn.config(state="normal")
            
            edit_entry.destroy()

        edit_entry.bind("<Return>", commit_change)
        edit_entry.bind("<Escape>", lambda e: edit_entry.destroy())
        # 失去焦点时自动尝试保存
        edit_entry.bind("<FocusOut>", lambda e: self.after(50, commit_change))

    def _save_data_changes(self):
        """保存所有挂起的数据修改，并在日志中详细记录生成的 SQL 及新旧值"""
        if not self.pending_updates: return
        
        if not messagebox.askyesno("确认", f"确定保存 {len(self.pending_updates)} 行修改？"):
            return
            
        sqls = []
        db_type = self.conn_data.get('db_type', 'MySQL')
        
        # 准备日志详情
        log_entries = []
        
        for item_id, updates in self.pending_updates.items():
            orig_row_vals = self.original_values[item_id]
            where_clauses = []
            where_params = []
            
            # 找到主键对应的值用于定位行
            for pk in self.current_query_pks:
                if pk in self.tree["columns"]:
                    pk_idx = list(self.tree["columns"]).index(pk)
                    where_clauses.append(f"`{pk}` = %s" if db_type == "MySQL" else f'"{pk}" = %s')
                    where_params.append(orig_row_vals[pk_idx])
            
            if not where_clauses: continue
            
            set_clauses = []
            set_params = []
            change_details = []
            
            for col, val in updates.items():
                col_idx = list(self.tree["columns"]).index(col)
                orig_val = orig_row_vals[col_idx]
                change_details.append(f"[{col}]: {orig_val} -> {val}")
                
                set_clauses.append(f"`{col}` = %s" if db_type == "MySQL" else f'"{col}" = %s')
                set_params.append(val)
            
            # 生成 SQL
            sql_text = f"UPDATE `{self.current_query_table}` SET {', '.join(set_clauses)} WHERE {' AND '.join(where_clauses)}"
            if db_type != "MySQL":
                # PostgreSQL: 使用 schema 限定表名
                qualified_table = f'"{self.schema_name}"."{self.current_query_table}"' if self.schema_name else f'"{self.current_query_table}"'
                sql_text = f'UPDATE {qualified_table} SET {", ".join(set_clauses)} WHERE {" AND ".join(where_clauses)}'
            
            params = set_params + where_params
            sqls.append((sql_text, params))
            
            # 存入待打印日志
            log_entries.append({
                "sql": sql_text,
                "params": params,
                "changes": change_details
            })
            
        def run():
            # 传递 schema 参数，execute_batch_sql 会在同一连接内先 SET search_path
            batch_schema = self.schema_name if db_type != "MySQL" else None
            success, msg = self.db_ops.execute_batch_sql(self.conn_data, sqls, database=self.db_name, schema=batch_schema)
            if success:
                def log_success():
                    self._log_msg(f"--- 批量更新执行成功 (表: {self.current_query_table}, 共 {len(sqls)} 条) ---")
                    for entry in log_entries:
                        self._log_msg(f"执行 SQL: {entry['sql']}")
                        self._log_msg(f"绑定参数: {entry['params']}")
                        self._log_msg(f"变更详情: {', '.join(entry['changes'])}")
                        self._log_msg("-" * 30)
                    self._run_query()
                self.after(0, log_success)
            else:
                self.after(0, lambda: messagebox.showerror("保存失败", msg))
        
        threading.Thread(target=run, daemon=True).start()

    def _cancel_data_changes(self):
        """取消所有挂起的数据修改"""
        if not self.pending_updates: return
        
        for item_id, orig_vals in self.original_values.items():
            self.tree.item(item_id, values=orig_vals, tags=())
            
        self.pending_updates.clear()
        self.original_values.clear()
        self.save_data_btn.config(state="disabled")
        self.cancel_data_btn.config(state="disabled")

    def _apply_theme(self):
        """应用主题样式并实现极致锐利的全网格分割线"""
        theme = self.settings.get('theme', '默认')
        font_family = self.settings.get('font_family', 'Microsoft YaHei')
        size_content = int(self.settings.get('font_size_content', 10))

        colors = get_theme_colors(theme)
        bg = colors["bg"]
        fg = colors["fg"]
        hb = colors["hb"]
        accent = colors["accent"]
        grid_color = colors["grid_color"]
        stripe_color = colors["stripe_color"]
        select_bg = colors["select_bg"]
        danger = colors["danger"]

        style = ttk.Style()
        try:
            # clam 引擎是目前 Tkinter 中网格线渲染最锐利的引擎
            if "clam" in style.theme_names():
                style.theme_use("clam")
        except: pass
            
        style.configure("TFrame", background=bg)
        style.configure("TLabel", background=bg, foreground=fg)
        
        # 动态计算行高以适配高 DPI（与 main_window 一致：2.2×，min 28）
        dynamic_row_height = int(size_content * 2.2)
        if dynamic_row_height < 28: dynamic_row_height = 28

        # 核心锐化配置
        style.configure("Treeview", 
                        background=bg, 
                        foreground=fg, 
                        fieldbackground=bg,
                        rowheight=dynamic_row_height,
                        gridcolor=grid_color, 
                        borderwidth=0,
                        font=(font_family, size_content))
        
        # 修正布局：确保 treearea 布局在绘制时不会出现子像素偏移导致的模糊
        style.layout("Treeview", [
            ('Treeview.treearea', {'sticky': 'nswe'})
        ])
        
        # 表头：强化边缘
        style.configure("Treeview.Heading", 
                        background=hb, 
                        foreground=fg, 
                        relief="ridge", 
                        borderwidth=1,
                        font=(font_family, size_content, "bold"))

        # 配置斑马纹颜色
        self.tree.tag_configure("odd", background=stripe_color)
        
        # 配置数据修改标记颜色（深色/浅色主题使用不同醒目色）
        if is_dark_theme(theme):
            modified_bg, modified_fg = "#553333", "#ff6666"
        else:
            modified_bg, modified_fg = "#fff3cd", "#856404"
        self.tree.tag_configure("modified", background=modified_bg, foreground=modified_fg)
        
        # 选中项
        style.map("Treeview", 
                  background=[('selected', select_bg)], 
                  foreground=[('selected', 'white')])
        
        # 更新 UI 子控件背景
        def update_bg(parent):
            for child in parent.winfo_children():
                try:
                    if isinstance(child, (tk.Frame, tk.Label)): 
                        child.configure(bg=bg)
                        if isinstance(child, tk.Label): child.configure(fg=fg)
                    if isinstance(child, (tk.Text, scrolledtext.ScrolledText)):
                        child.configure(bg=bg, fg=fg, insertbackground=fg, font=(font_family, size_content))
                    if isinstance(child, tk.Entry):
                        child.configure(bg=bg, fg=fg, insertbackground=fg, font=(font_family, size_content))
                except: pass
                update_bg(child)
        
        update_bg(self)
        self.editor.configure(bg=bg, fg=fg, insertbackground=fg, font=(font_family, size_content))
        self.msg_area.configure(bg=bg, fg=fg, font=(font_family, size_content))
        
        # 更新 footer、find bar、msg 工具栏等特殊区域
        # 这些区域使用 hb 背景而非 bg，需同时更新其子控件
        for frame_attr in ['toolbar', 'data_footer', 'find_frame', 'msg_toolbar']:
            frame = getattr(self, frame_attr, None)
            if frame and frame.winfo_exists():
                frame.configure(bg=hb)
                for child in frame.winfo_children():
                    try:
                        if isinstance(child, (tk.Label, tk.Frame)):
                            child.configure(bg=hb)
                        elif isinstance(child, tk.Entry):
                            child.configure(bg=bg, fg=fg, insertbackground=fg)
                    except: pass
        if hasattr(self, 'edit_tip_label') and self.edit_tip_label.winfo_exists():
            self.edit_tip_label.configure(fg="#888" if not is_dark_theme(theme) else "#aaa")
        
        self._highlight_syntax()
        # 将主题同步到 AI 聊天面板
        self.after(50, lambda: self._propagate_theme_to_chat())

    def _propagate_theme_to_chat(self):
        """将当前主题同步到 AI 聊天面板"""
        try:
            if hasattr(self, 'chat_container'):
                for child in self.chat_container.winfo_children():
                    if hasattr(child, '_apply_theme'):
                        child._apply_theme()
                        break
        except:
            pass

    def _apply_settings(self):
        """MainWindow 调用此方法以同步设置"""
        self._apply_theme()

    def _on_paste(self, event=None):
        """异步分块粘贴，处理超大文本防止锁死 (优化版：增加吞吐量并禁用换行)"""
        try:
            text = self.clipboard_get()
        except:
            return
            
        if not text: return
        
        # 阈值：超过 50,000 字符启动异步模式
        if len(text) < 50000:
            return # 使用原生粘贴
            
        # 拦截原生粘贴
        progress = ProgressDialog(self, message="正在极速分块粘贴超大脚本...")
        
        # 性能极致优化：暂时关闭 undo 和 wrap 以提升插入速度
        orig_undo = self.editor.cget("undo")
        orig_wrap = self.editor.cget("wrap")
        self.editor.configure(undo=False, wrap="none")
        
        # 加大块大小：从 20k 增加到 150k
        chunk_size = 150000 
        total_len = len(text)
        
        def insert_chunk(start):
            if not self.winfo_exists(): return
            end = min(start + chunk_size, total_len)
            self.editor.insert(tk.INSERT, text[start:end])
            
            # 更新进度
            pct = int((end / total_len) * 100)
            progress.update_progress(pct, f"已粘贴 {pct}% ({(end / 1024 / 1024):.1f} MB / {(total_len / 1024 / 1024):.1f} MB)")
            
            if end < total_len:
                # 缩短延迟，1ms 仅用于释放 UI 线程
                self.after(1, lambda: insert_chunk(end))
            else:
                progress.close()
                self.editor.configure(undo=orig_undo, wrap=orig_wrap)
                self.editor.edit_separator()
                self._log_msg(f"超大脚本粘贴完成 (共 {(total_len / 1024 / 1024):.2f} MB)")
                self._highlight_syntax()
                
        insert_chunk(0)
        return "break" # 阻止 Tkinter 默认粘贴行为

    def _highlight_syntax(self):
        """SQL 语法高亮 (增加超大文本降级保护)"""
        content = self.editor.get("1.0", tk.END)
        
        # 极致降级：超过 100,000 字符直接跳过渲染
        if len(content) > 100000:
            for tag in ["keyword", "string", "comment"]:
                self.editor.tag_remove(tag, "1.0", tk.END)
            # 如果是刚刚触发的超长粘贴，给出一次提示
            if not hasattr(self, '_highlight_warn_done') or not self._highlight_warn_done:
                self._log_msg("文本过大 (已超过 100KB)，已自动关闭语法高亮以保证编辑性能", "WARN")
                self._highlight_warn_done = True
            return

        self._highlight_warn_done = False
        for tag in ["keyword", "string", "comment"]:
            self.editor.tag_remove(tag, "1.0", tk.END)
            
        keywords = r'\b(SELECT|FROM|WHERE|INSERT|UPDATE|DELETE|JOIN|LEFT|RIGHT|INNER|ON|GROUP|BY|ORDER|HAVING|LIMIT|CREATE|TABLE|ALTER|DROP|INDEX|VIEW|FUNCTION|PROCEDURE|BEGIN|END|IF|ELSE|CASE|WHEN|THEN|AND|OR|NOT|IN|IS|NULL|LIKE|AS|SET|VALUES|INTO|COMMIT|ROLLBACK|USE|DATABASES|SHOW|TRUNCATE)\b'
        for match in re.finditer(keywords, content, re.IGNORECASE):
            self.editor.tag_add("keyword", f"1.0+{match.start()}c", f"1.0+{match.end()}c")
            
        for match in re.finditer(r"'[^']*'|\"[^\"]*\"", content):
            self.editor.tag_add("string", f"1.0+{match.start()}c", f"1.0+{match.end()}c")
            
        for match in re.finditer(r"--.*$|#.*$|/\*[\s\S]*?\*/", content, re.MULTILINE):
            self.editor.tag_add("comment", f"1.0+{match.start()}c", f"1.0+{match.end()}c")

        # 使用主题定义的语法高亮色
        colors = get_theme_colors(self.settings.get('theme', '默认'))
        is_dark = is_dark_theme(self.settings.get('theme', '默认'))
        keyword_fg = colors["syntax_keyword"]
        string_fg = colors["syntax_string"]
        comment_fg = colors["syntax_comment"]
        # 深色主题不使用 bold 以避免模糊，浅色主题使用 bold 增强可读性
        kw_font = (self.settings.get('font_family', 'Consolas'), 12, "bold") if not is_dark else (self.settings.get('font_family', 'Consolas'), 12)
        self.editor.tag_configure("keyword", foreground=keyword_fg, font=kw_font)
        self.editor.tag_configure("string", foreground=string_fg)
        self.editor.tag_configure("comment", foreground=comment_fg)

    def _apply_completion(self):
        """应用选中的代码补全 - Navicat 风格
        - 函数自动补充括号 ()
        - 列名补全后可选追加逗号
        - 智能光标定位
        """
        if not self.comp_listbox:
            return
        sel = self.comp_listbox.curselection()
        if not sel:
            return

        # 获取选中的补全项
        if hasattr(self, '_completion_items') and self._completion_items:
            raw = self._completion_items[sel[0]]
        else:
            raw = self.comp_listbox.get(sel[0])
            if '  -- ' in raw:
                raw = raw.split('  -- ')[0]

        # 提取插入文本
        insert_text = None
        is_function = False
        if isinstance(raw, CompletionItem):
            insert_text = raw.insert_text
            is_function = (raw.type == CompletionItem.TYPE_FUNCTION)
        elif isinstance(raw, tuple):
            insert_text = raw[0]
        else:
            insert_text = raw

        idx = self.editor.index(tk.INSERT)
        line_start = f"{idx} linestart"
        line_content = self.editor.get(line_start, idx)

        # 删除已输入的前缀
        match = re.search(r'[\w\.\*]+$', line_content)
        if match:
            start_pos = f"{line_start}+{match.start()}c"
            self.editor.delete(start_pos, idx)
            self.editor.insert(start_pos, insert_text)
            new_cursor_pos = f"{line_start}+{match.start() + len(insert_text)}c"
        else:
            self.editor.insert(idx, insert_text)
            new_cursor_pos = f"{idx}+{len(insert_text)}c"

        # 函数补全：光标定位到括号内
        if is_function and insert_text.endswith('('):
            # 插入闭合括号
            self.editor.insert(new_cursor_pos, ")")
            # 光标定位到括号内
            inner_pos = f"{line_start}+{len(insert_text) - 1}c" if match else f"{idx}+{len(insert_text) - 1}c"
            self.editor.mark_set(tk.INSERT, inner_pos)
        else:
            self.editor.mark_set(tk.INSERT, new_cursor_pos)

        self._hide_completions()
        self._highlight_syntax()

        # 补全后 50ms 短暂检查是否需要继续补全（如别名.的后续）
        self.after(50, lambda: self._show_completions(manual=False))

    def _run_sql_file(self):
        """弹出文件选择框，流式执行大型 SQL 文件"""
        path = filedialog.askopenfilename(filetypes=[("SQL Files", "*.sql"), ("All Files", "*.*")])
        if not path:
            return
            
        file_size_mb = os.path.getsize(path) / 1024 / 1024
        if not messagebox.askokcancel("执行 SQL 文件", f"确定执行以下文件吗？\n文件: {os.path.basename(path)}\n大小: {file_size_mb:.2f} MB\n\n警告：执行期间无法撤回，完成后请手动刷新数据。"):
            return

        self.is_running = True
        
        self.exec_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        
        self._log_msg(f"开始流式执行 SQL 文件: {path} (大小: {file_size_mb:.2f} MB)")
        
        progress = ProgressDialog(self, message="正在准备流式解析...")
        
        def run():
            try:
                if self.cancel_event: self.cancel_event.clear()
                
                def on_progress(pct, msg):
                    self.after(0, lambda: progress.update_progress(pct, msg))

                schema = self.schema_name if self.conn_data.get('db_type') == 'PostgreSQL' else None
                success, stmt_count = self.db_ops.execute_sql_from_file(
                    self.conn_data, path, database=self.db_name,
                    cancel_event=self.cancel_event, schema=schema,
                    progress_callback=on_progress
                )
                self.after(0, lambda: self._on_file_execute_done(progress, success, stmt_count, None))
            except Exception as e:
                self.after(0, lambda e=e: self._on_file_execute_done(progress, False, 0, str(e)))
        
        threading.Thread(target=run, daemon=True).start()

    def _on_file_execute_done(self, progress, success, stmt_count, error):
        progress.close()
        self.is_running = False
             
        self.exec_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        
        if error:
            self._log_msg(f"文件执行失败: {error}", "ERROR")
            messagebox.showerror("执行失败", f"文件执行中断：\n{error}")
        else:
            self._log_msg(f"文件执行完成！共执行 {stmt_count} 条有效 SQL 语句。", "INFO")
            messagebox.showinfo("成功", f"SQL 文件执行完成！\n共执行语句：{stmt_count} 条")
        self.result_notebook.select(1)

    def _load_persistent_logs(self):
        """从数据库恢复历史日志 (按路径)"""
        try:
            content = self.storage.get_workbench_log(self.script_path)
            if content:
                self.msg_area.configure(state="normal")
                self.msg_area.delete("1.0", tk.END)
                self.msg_area.insert("1.0", content)
                self.msg_area.see(tk.END)
                self.msg_area.configure(state="disabled")
        except Exception as e:
            print(f"加载历史日志失败: {e}")
