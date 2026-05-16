from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.backends import default_backend
import binascii
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import base64
import xml.etree.ElementTree as ET
from xml.dom import minidom
import os
import logging
import re
from datetime import datetime

# 彻底静默 fontTools 的日志输出，解决 MERG NOT subset 等干扰信息
logging.getLogger("fontTools").setLevel(logging.ERROR)

# ── Mermaid.js 内嵌缓存 ──
_MERMAID_JS_CACHE = None

def _get_mermaid_js():
    """下载 mermaid.min.js 并缓存，返回 JS 内容；失败返回 None"""
    global _MERMAID_JS_CACHE
    if _MERMAID_JS_CACHE is not None:
        return _MERMAID_JS_CACHE
    try:
        import urllib.request
        url = "https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.min.js"
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            _MERMAID_JS_CACHE = resp.read().decode('utf-8')
        return _MERMAID_JS_CACHE
    except Exception as e:
        logging.warning(f"Mermaid.js 下载失败，回退 CDN: {e}")
        return None

class NavicatCipher:
    """针对 Navicat 12+ (NCX 1.5) 的密码加密工具类 (AES-128-CBC)"""
    def __init__(self):
        self.key = b'libcckeylibcckey'
        self.iv = b'libcciv libcciv '

    def encrypt(self, plaintext: str) -> str:
        if not plaintext:
            return ""
        try:
            # PKCS7 填充
            padder = padding.PKCS7(128).padder()
            padded_data = padder.update(plaintext.encode('utf-8')) + padder.finalize()
            
            # AES-128-CBC 加密
            cipher = Cipher(algorithms.AES(self.key), modes.CBC(self.iv), backend=default_backend())
            encryptor = cipher.encryptor()
            ct = encryptor.update(padded_data) + encryptor.finalize()
            
            # 返回大写十六进制字符串
            return ct.hex().upper()
        except Exception:
            return ""

class Exporter:
    @staticmethod
    def _clean_value(val):
        """
        清理数据值，防止 openpyxl 报 IllegalCharacterError。
        1. 处理字节流 (解决 BIT 字段显示为 b'\\x00' 的问题)
        2. 剔除所有 XML 禁止的控制字符
        """
        if val is None: return ""
        
        # 处理字节流
        if isinstance(val, bytes):
            if val == b'\x00': return 0
            if val == b'\x01': return 1
            try:
                val = val.decode('utf-8')
            except:
                val = str(val)

        s_val = str(val)
        # 精准剔除 XML 禁止的 ASCII 控制字符 (0-31)，保留 \n \r \t (9, 10, 13)
        # 匹配: 0-8, 11-12, 14-31
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s_val)

    @staticmethod
    def _sanitize_sheet_name(name):
        """清洗并截断 Excel 工作表名称，确保符合 Excel 规范 (最长 31 字符, 无非法字符)"""
        if not name: return "Sheet"
        # 1. 剔除非法字符 \ / ? * [ ] :
        invalid_chars = r'[\\/\?\*\[\]:]'
        clean_name = re.sub(invalid_chars, '_', str(name))
        # 2. 彻底清理控制字符
        clean_name = Exporter._clean_value(clean_name)
        # 3. 截断至 31 个字符并去除前后空格
        return clean_name[:31].strip()

    @staticmethod
    def export_table_structure_to_html(structures, file_path):
        """将表结构导出为带模糊搜索功能的 HTML 网页"""
        html_template = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>数据库表结构报告</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; margin: 0; display: flex; height: 100vh; overflow: hidden; }
        #sidebar { width: 300px; border-right: 1px solid #ddd; display: flex; flex-direction: column; background: #f8f9fa; }
        #search-box { padding: 15px; border-bottom: 1px solid #ddd; }
        #search-input { width: 100%; padding: 8px; box-sizing: border-box; border: 1px solid #ccc; border-radius: 4px; }
        #table-list { flex: 1; overflow-y: auto; list-style: none; padding: 0; margin: 0; }
        #table-list li { padding: 10px 15px; cursor: pointer; border-bottom: 1px solid #eee; transition: background 0.2s; }
        #table-list li:hover { background: #e9ecef; }
        #table-list li.active { background: #007bff; color: white; }
        #content { flex: 1; overflow-y: auto; padding: 20px; background: white; }
        .table-detail { display: none; }
        .table-detail.active { display: block; }
        table { width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 14px; }
        th, td { border: 1px solid #dee2e6; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; font-weight: bold; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        .table-info { background: #e7f3ff; padding: 10px; border-radius: 4px; border-left: 4px solid #007bff; margin-bottom: 20px; }
        h2 { margin-top: 0; }
    </style>
</head>
<body>
    <div id="sidebar">
        <div id="search-box">
            <input type="text" id="search-input" placeholder="模糊搜索表名..." oninput="filterTables()">
        </div>
        <ul id="table-list">
            {% TABLE_LIST_ITEMS %}
        </ul>
    </div>
    <div id="content">
        <div id="welcome" style="text-align:center; padding-top:100px; color:#666;">
            <h1>请从左侧选择一个表查看详情</h1>
        </div>
        {% TABLE_DETAILS %}
    </div>

    <script>
        function filterTables() {
            const input = document.getElementById('search-input');
            const filter = input.value.toLowerCase();
            const li = document.getElementById('table-list').getElementsByTagName('li');
            for (let i = 0; i < li.length; i++) {
                const txtValue = li[i].textContent || li[i].innerText;
                if (txtValue.toLowerCase().indexOf(filter) > -1) {
                    li[i].style.display = "";
                } else {
                    li[i].style.display = "none";
                }
            }
        }

        function showTable(tableName) {
            document.getElementById('welcome').style.display = 'none';
            // 切换列表状态
            const listItems = document.querySelectorAll('#table-list li');
            listItems.forEach(li => {
                if(li.getAttribute('onclick').includes("'" + tableName + "'")) {
                    li.classList.add('active');
                } else {
                    li.classList.remove('active');
                }
            });
            // 切换内容
            const details = document.querySelectorAll('.table-detail');
            details.forEach(d => {
                if(d.id === 'detail-' + tableName) {
                    d.classList.add('active');
                } else {
                    d.classList.remove('active');
                }
            });
            // 滚动到顶部
            document.getElementById('content').scrollTop = 0;
        }
    </script>
</body>
</html>
"""
        list_items = []
        details = []
        
        # 排序表名
        sorted_tables = sorted(structures.keys())
        
        for table_name in sorted_tables:
            info = structures[table_name]
            # 生成侧边栏列表项
            list_items.append(f'<li onclick="showTable(\'{table_name}\')">{table_name}</li>')
            
            # 生成详情 HTML
            rows = []
            for col in info['columns']:
                rows.append(f"""
                <tr>
                    <td>{col['Field']}</td>
                    <td>{col['Type']}</td>
                    <td>{col.get('Collation', '') or ''}</td>
                    <td>{col['Null']}</td>
                    <td>{col['Key']}</td>
                    <td>{col['Default'] or ''}</td>
                    <td>{col['Extra']}</td>
                    <td>{col['Comment']}</td>
                </tr>
                """)
            
            detail_html = f"""
            <div class="table-detail" id="detail-{table_name}">
                <div class="table-info">
                    <h2>表名: {table_name}</h2>
                    <p><strong>备注:</strong> {info['comment'] or '无'}</p>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>字段</th>
                            <th>类型</th>
                            <th>字符集</th>
                            <th>允许空</th>
                            <th>键</th>
                            <th>默认值</th>
                            <th>额外</th>
                            <th>备注</th>
                        </tr>
                    </thead>
                    <tbody>
                        {''.join(rows)}
                    </tbody>
                </table>
            </div>
            """
            details.append(detail_html)

        full_html = html_template.replace('{% TABLE_LIST_ITEMS %}', '\n'.join(list_items))
        full_html = full_html.replace('{% TABLE_DETAILS %}', '\n'.join(details))
        
        with open(file_path, "w", encoding='utf-8') as f:
            f.write(full_html)

    @staticmethod
    def export_er_to_markdown(structures, relations, file_path):
        """将数据库结构与关联关系导出为 Markdown 文件 (含 Mermaid 源码)"""
        
        def sanitize(text):
            if not text: return "unknown"
            cleaned = re.sub(r'[^a-zA-Z0-9_]', '_', str(text))
            if cleaned and cleaned[0].isdigit():
                cleaned = "n_" + cleaned
            return cleaned

        lines = ["# 数据库 ER 关系报告\n"]
        lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # 1. Mermaid 可视化源码 (部分编辑器如 Typora/Obsidian 支持直接渲染)
        lines.append("## 1. 可视化图表 (Mermaid)")
        lines.append("```mermaid")
        lines.append("erDiagram")
        for table_name, info in sorted(structures.items()):
            safe_table = sanitize(table_name)
            lines.append(f"    {safe_table} {{")
            for col in info['columns']:
                raw_type = col['Type'].split('(')[0]
                lines.append(f"        {sanitize(raw_type)} {sanitize(col['Field'])}")
            lines.append("    }")
        
        for rel in relations:
            t1 = sanitize(rel['table'])
            t2 = sanitize(rel['ref_table'])
            lines.append(f"    {t2} ||--o{{ {t1} : \"{sanitize(rel['column'])}\"")
        lines.append("```\n")

        # 2. 关系明细表
        lines.append("## 2. 关联关系明细")
        lines.append("| 子表 (从表) | 外键字段 | 指向 | 父表 (主表) | 关联主键 |")
        lines.append("| --- | --- | --- | --- | --- |")
        for r in relations:
            lines.append(f"| {r['table']} | {r['column']} | -> | {r['ref_table']} | {r['ref_column']} |")
        
        lines.append("\n---\n")
        
        # 3. 数据字典
        lines.append("## 3. 数据字典 (表结构)")
        for table_name, info in sorted(structures.items()):
            lines.append(f"### {table_name}")
            lines.append(f"**备注:** {info['comment'] or '无'}\n")
            lines.append("| 字段 | 类型 | 为空 | 键 | 备注 |")
            lines.append("| --- | --- | --- | --- | --- |")
            for c in info['columns']:
                lines.append(f"| {c['Field']} | {c['Type']} | {c['Null']} | {c['Key']} | {c['Comment']} |")
            lines.append("\n")

        with open(file_path, "w", encoding='utf-8') as f:
            f.write("\n".join(lines))

    @staticmethod
    def export_er_to_pdf(structures, relations, file_path):
        """将数据库结构与关联关系导出为 PDF 报告"""
        from fpdf import FPDF
        import warnings
        logging.getLogger("fontTools").setLevel(logging.ERROR)
        
        class PDF(FPDF):
            def header(self):
                self.set_font('chinese', 'B', 15)
                self.cell(0, 10, '数据库 ER 关系报告', 0, 1, 'C')
                self.ln(5)

            def footer(self):
                self.set_y(-15)
                self.set_font('chinese', 'I', 8)
                self.cell(0, 10, f'第 {self.page_no()} 页', 0, 0, 'C')

        pdf = PDF()
        font_paths = ["C:\\Windows\\Fonts\\msyh.ttc", "C:\\Windows\\Fonts\\simhei.ttf"]
        for path in font_paths:
            if os.path.exists(path):
                pdf.add_font('chinese', '', path)
                pdf.add_font('chinese', 'B', path)
                pdf.add_font('chinese', 'I', path)
                break
        
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        
        # 1. 关联关系部分
        pdf.set_font('chinese', 'B', 14)
        pdf.cell(0, 10, "一、关联关系图谱 (外键引用)", 0, 1, 'L')
        pdf.ln(2)
        
        rel_headers = ["从表 (引用者)", "外键字段", "主表 (被引用)", "关联主键"]
        rel_widths = [55, 40, 55, 40]
        
        with pdf.table(col_widths=rel_widths, borders_layout="ALL", line_height=7) as table:
            pdf.set_font('chinese', 'B', 10)
            pdf.set_fill_color(240, 240, 240)
            h_row = table.row()
            for h in rel_headers: h_row.cell(h)
            
            pdf.set_font('chinese', '', 9)
            for r in relations:
                row = table.row()
                row.cell(r['table']); row.cell(r['column']); row.cell(r['ref_table']); row.cell(r['ref_column'])
        
        pdf.ln(10)

        # 2. 数据字典部分
        pdf.set_font('chinese', 'B', 14)
        pdf.cell(0, 10, "二、数据字典 (详细结构)", 0, 1, 'L')
        pdf.ln(5)
        
        for table_name, info in sorted(structures.items()):
            if pdf.get_y() > 240: pdf.add_page()
            
            pdf.set_font('chinese', 'B', 11)
            pdf.set_fill_color(220, 230, 240)
            pdf.cell(0, 8, f"表名: {table_name}", 1, 1, 'L', fill=True)
            pdf.set_font('chinese', '', 9)
            pdf.multi_cell(0, 7, f"备注: {info['comment'] or '无'}", border=1)
            
            headers = ["字段", "类型", "为空", "键", "备注"]
            widths = [40, 45, 15, 10, 80]
            with pdf.table(col_widths=widths, borders_layout="ALL", line_height=6) as table:
                pdf.set_font('chinese', 'B', 9)
                pdf.set_fill_color(245, 245, 245)
                hr = table.row()
                for h in headers: hr.cell(h)
                
                pdf.set_font('chinese', '', 8)
                for c in info['columns']:
                    tr = table.row()
                    tr.cell(c['Field']); tr.cell(c['Type']); tr.cell(c['Null']); tr.cell(c['Key']); tr.cell(c['Comment'])
            pdf.ln(8)

        pdf.output(file_path)

    @staticmethod
    def export_er_to_excel(structures, relations, file_path):
        """将数据库结构与关联关系导出为高质量的 Excel 文件 (ER 表格式)"""
        wb = openpyxl.Workbook()
        
        # --- Sheet 1: 数据字典 (含结构) ---
        ws_dict = wb.active
        ws_dict.title = Exporter._sanitize_sheet_name("1. 数据字典 (表结构详情)")
        
        # 样式定义
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        curr_row = 1
        for table_name, info in sorted(structures.items()):
            # 表大标题
            ws_dict.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
            # 清理数据内容中的控制字符
            safe_title = Exporter._clean_value(f"数据表: {table_name}  (备注: {info['comment'] or '无'})")
            title_cell = ws_dict.cell(row=curr_row, column=1, value=safe_title)
            title_cell.fill = header_fill
            title_cell.font = header_font
            curr_row += 1
            
            # 列标题
            col_headers = ["字段名", "数据类型", "允许为空", "主/外键", "默认值", "字段备注"]
            ws_dict.append([Exporter._clean_value(h) for h in col_headers])
            for i in range(1, 7):
                ws_dict.cell(row=curr_row, column=i).fill = sub_header_fill
                ws_dict.cell(row=curr_row, column=i).font = Font(bold=True)
            curr_row += 1
            
            # 写入字段
            for c in info['columns']:
                row_data = [
                    c['Field'], c['Type'], c['Null'], c['Key'], 
                    c['Default'] if c['Default'] is not None else '', 
                    c['Comment']
                ]
                ws_dict.append([Exporter._clean_value(v) for v in row_data])
                curr_row += 1
            
            ws_dict.append([]) # 间隔空行
            curr_row += 1

        for i, w in enumerate([30, 20, 10, 10, 15, 50], 1):
            ws_dict.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # --- Sheet 2: 关联关系矩阵 ---
        ws_rel = wb.create_sheet(title=Exporter._sanitize_sheet_name("2. 关联关系图谱 (外键引用)"))
        rel_header_fill = PatternFill(start_color="76933C", end_color="76933C", fill_type="solid")
        
        rel_headers = ["从表 (引用者)", "外键字段", "指向 ->", "主表 (被引用)", "关联主键", "完整关系描述"]
        ws_rel.append([Exporter._clean_value(h) for h in rel_headers])
        for i, cell in enumerate(ws_rel[1], 1):
            cell.fill = rel_header_fill
            cell.font = header_font
            ws_rel.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 30

        for r in relations:
            row_data = [
                r['table'], r['column'], "关联到", 
                r['ref_table'], r['ref_column'],
                f"{r['table']}.{r['column']}  ===>  {r['ref_table']}.{r['ref_column']}"
            ]
            ws_rel.append([Exporter._clean_value(v) for v in row_data])

        wb.save(file_path)

    @staticmethod
    def export_er_to_html(structures, relations, file_path):
        """将数据库结构与关联关系导出为带侧边栏搜索和表选择功能的交互式 ER 图"""
        import json
        
        def sanitize(text):
            if not text: return "unknown"
            cleaned = re.sub(r'[^a-zA-Z0-9_]', '_', str(text))
            if cleaned and cleaned[0].isdigit():
                cleaned = "n_" + cleaned
            return cleaned

        # 准备 JSON 数据用于前端搜索
        entities_data = []
        for table_name, info in sorted(structures.items()):
            fields = []
            for col in info['columns']:
                fields.append({
                    "name": col['Field'],
                    "type": sanitize(col['Type'].split('(')[0])
                })
            entities_data.append({
                "id": sanitize(table_name),
                "display": table_name,
                "fields": fields
            })
            
        links_data = []
        for rel in relations:
            links_data.append({
                "source": sanitize(rel['ref_table']),
                "target": sanitize(rel['table']),
                "label": rel['column']
            })

        # 尝试下载并内嵌 mermaid.js（自包含，避免 file:// CORS 问题）
        mermaid_js = _get_mermaid_js()
        if mermaid_js:
            mermaid_script = f"<script>\n{mermaid_js}\n</script>"
        else:
            mermaid_script = '<script src="https://cdn.jsdelivr.net/npm/mermaid/dist/mermaid.min.js"></script>'

        html_template = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>数据库 ER 关系图 - 交互式查看器</title>
    <MERMAID_SCRIPT>
    <style>
        body, html {{ 
            margin: 0; padding: 0; width: 100%; height: 100%; 
            font-family: 'Segoe UI', Tahoma, sans-serif;
            display: flex; overflow: hidden; background: #f0f2f5;
        }}
        #sidebar {{
            width: 320px; height: 100%; background: white;
            border-right: 1px solid #ddd; display: flex; flex-direction: column;
            z-index: 1000; box-shadow: 2px 0 10px rgba(0,0,0,0.05);
        }}
        .sidebar-header {{ padding: 20px; border-bottom: 1px solid #eee; }}
        .sidebar-header h1 {{ margin: 0 0 15px 0; font-size: 18px; color: #333; }}
        
        .search-box {{ position: relative; }}
        input {{
            width: 100%; padding: 10px 12px; border: 1px solid #ddd;
            border-radius: 6px; font-size: 14px; outline: none; transition: border 0.2s;
        }}
        input:focus {{ border-color: #007bff; }}
        
        .action-bar {{ display: flex; gap: 10px; margin-top: 15px; }}
        .btn {{
            flex: 1; padding: 6px; font-size: 12px; cursor: pointer;
            background: #f8f9fa; border: 1px solid #ddd; border-radius: 4px;
        }}
        .btn:hover {{ background: #e9ecef; }}
        .btn-primary {{ background: #007bff; color: white; border-color: #007bff; }}
        .btn-primary:hover {{ background: #0069d9; }}
        
        #table-list-container {{ flex: 1; overflow-y: auto; padding: 10px 20px; }}
        .table-item {{
            display: flex; align-items: center; padding: 8px 0;
            border-bottom: 1px solid #f9f9f9; font-size: 13px; cursor: pointer;
        }}
        .table-item:hover {{ background: #fcfcfc; }}
        .table-item input {{ width: auto; margin-right: 10px; cursor: pointer; }}
        .table-item label {{ flex: 1; cursor: pointer; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        
        #main-area {{ flex: 1; position: relative; overflow: hidden; }}
        #viewport {{ width: 100%; height: 100%; cursor: grab; display: flex; align-items: center; justify-content: center; }}
        #viewport:active {{ cursor: grabbing; }}
        #diagram-wrapper {{ transform-origin: 0 0; transition: transform 0.1s; }}
        
        .floating-info {{
            position: absolute; bottom: 20px; right: 20px;
            background: rgba(255,255,255,0.9); padding: 10px;
            border-radius: 6px; font-size: 11px; color: #666; pointer-events: none;
        }}
    </style>
</head>
<body>
    <div id="sidebar">
        <div class="sidebar-header">
            <h1>📊 数据库 ER 图</h1>
            <div class="search-box">
                <input type="text" id="search-input" placeholder="搜索数据表..." oninput="filterList()">
            </div>
            <div class="action-bar">
                <button class="btn" onclick="setAll(true)">全选</button>
                <button class="btn" onclick="setAll(false)">全不选</button>
                <button class="btn" id="compactBtn" onclick="toggleCompact()">📐 紧凑</button>
                <button class="btn btn-primary" onclick="resetView()">重置视图</button>
            </div>
        </div>
        <div id="table-list-container">
            <!-- 列表动态生成 -->
        </div>
    </div>

    <div id="main-area">
        <div id="viewport">
            <div id="diagram-wrapper">
                <pre class="mermaid" id="mermaid-pre"></pre>
            </div>
        </div>
        <div class="floating-info">
            滚轮: 缩放 | 拖拽: 移动 | 仅渲染选中的表及关系<br>
            生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
    </div>

    <script>
        const entities = {json.dumps(entities_data)};
        const links = {json.dumps(links_data)};
        const selectedIds = new Set(entities.map(e => e.id)); // 初始全选
        const COMPACT_THRESHOLD = 50;  // 超过此数默认紧凑模式
        let compactMode = entities.length > COMPACT_THRESHOLD;
        let isRendering = false;
        let renderQueue = [];

        function initList() {{
            const container = document.getElementById('table-list-container');
            container.innerHTML = '';
            entities.forEach(e => {{
                const div = document.createElement('div');
                div.className = 'table-item';
                div.id = 'item-' + e.id;
                div.innerHTML = `
                    <input type="checkbox" id="chk-${{e.id}}" ${{selectedIds.has(e.id) ? 'checked' : ''}} 
                           onchange="toggleTable('${{e.id}}')">
                    <label for="chk-${{e.id}}" title="${{e.display}}">${{e.display}}</label>
                `;
                container.appendChild(div);
            }});
            // 紧凑模式提示
            if (compactMode) {{
                const banner = document.createElement('div');
                banner.id = 'compact-banner';
                banner.style.cssText = 'padding:8px 12px;margin:0 10px 8px;background:#fff3cd;border-radius:4px;font-size:12px;color:#856404;';
                banner.textContent = `📐 紧凑模式 (${{entities.length}} 张表)，点击「完整」切换显示字段`;
                container.parentNode.insertBefore(banner, container);
            }}
            updateCompactBtn();
        }}

        function filterList() {{
            const term = document.getElementById('search-input').value.toLowerCase();
            entities.forEach(e => {{
                const el = document.getElementById('item-' + e.id);
                if (e.display.toLowerCase().includes(term)) {{
                    el.style.display = 'flex';
                }} else {{
                    el.style.display = 'none';
                }}
            }});
        }}

        function setAll(val) {{
            entities.forEach(e => {{
                const chk = document.getElementById('chk-' + e.id);
                chk.checked = val;
                if (val) selectedIds.add(e.id);
                else selectedIds.delete(e.id);
            }});
            startRender();
        }}

        function toggleTable(id) {{
            const chk = document.getElementById('chk-' + id);
            if (chk.checked) selectedIds.add(id);
            else selectedIds.delete(id);
            debounceRender();
        }}

        function toggleCompact() {{
            compactMode = !compactMode;
            updateCompactBtn();
            const banner = document.getElementById('compact-banner');
            if (banner) banner.remove();
            if (compactMode && entities.length > COMPACT_THRESHOLD) {{
                const container = document.getElementById('table-list-container');
                const banner2 = document.createElement('div');
                banner2.id = 'compact-banner';
                banner2.style.cssText = 'padding:8px 12px;margin:0 10px 8px;background:#fff3cd;border-radius:4px;font-size:12px;color:#856404;';
                banner2.textContent = `📐 紧凑模式，点击「完整」切换显示字段`;
                container.parentNode.insertBefore(banner2, container);
            }}
            startRender();
        }}

        function updateCompactBtn() {{
            const btn = document.getElementById('compactBtn');
            if (btn) btn.textContent = compactMode ? '📐 完整' : '📐 紧凑';
        }}

        let renderTimer;
        function debounceRender() {{
            clearTimeout(renderTimer);
            renderTimer = setTimeout(startRender, 300);
        }}

        // ── 分块渲染引擎 ──
        const CHUNK_SIZE = 25;

        function startRender() {{
            if (isRendering) {{
                renderQueue.push(true);
                return;
            }}
            isRendering = true;
            renderQueue = [];
            showProgress(true);
            renderChunk(0);
        }}

        function showProgress(show) {{
            let el = document.getElementById('render-progress');
            if (!show) {{
                if (el) el.remove();
                return;
            }}
            if (!el) {{
                el = document.createElement('div');
                el.id = 'render-progress';
                el.style.cssText = 'position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);background:rgba(0,0,0,0.7);color:#fff;padding:12px 24px;border-radius:8px;font-size:14px;z-index:999;pointer-events:none;';
                document.getElementById('main-area').appendChild(el);
            }}
            el.textContent = '⏳ 渲染中...';
        }}

        function renderChunk(startIdx) {{
            const activeEntities = entities.filter(e => selectedIds.has(e.id));
            if (activeEntities.length === 0) {{
                document.getElementById('diagram-wrapper').innerHTML = '<div style="color:#999;text-align:center;margin-top:200px;">请在左侧勾选要查看的表</div>';
                showProgress(false);
                isRendering = false;
                return;
            }}

            const chunk = activeEntities.slice(startIdx, startIdx + CHUNK_SIZE);
            if (chunk.length === 0) {{
                finalizeAndRender(activeEntities);
                return;
            }}

            let source = startIdx === 0 ? `erDiagram\n` : '';
            chunk.forEach(e => {{
                if (compactMode) {{
                    source += `    ${{e.id}}[${{e.display}}]\n`;
                }} else {{
                    source += `    ${{e.id}} {{\n`;
                    e.fields.forEach(f => {{
                        source += `        ${{f.type}} ${{f.name}}\n`;
                    }});
                    source += "    }}\\n";
                }}
            }});

            const wrapper = document.getElementById('diagram-wrapper');
            if (startIdx === 0) {{
                wrapper.innerHTML = `<pre class="mermaid">${{source}}</pre>`;
            }} else {{
                const pre = wrapper.querySelector('pre.mermaid');
                if (pre) pre.textContent += '\\n' + source;
            }}

            const progressEl = document.getElementById('render-progress');
            if (progressEl) {{
                const pct = Math.min(99, Math.round((startIdx + chunk.length) / activeEntities.length * 100));
                progressEl.textContent = `⏳ 渲染中... ${{pct}}% (${{Math.min(startIdx + CHUNK_SIZE, activeEntities.length)}}/${{activeEntities.length}})`;
            }}

            requestAnimationFrame(() => renderChunk(startIdx + CHUNK_SIZE));
        }}

        function finalizeAndRender(activeEntities) {{
            const wrapper = document.getElementById('diagram-wrapper');
            const pre = wrapper.querySelector('pre.mermaid');
            if (!pre) {{
                showProgress(false);
                isRendering = false;
                return;
            }}

            let relSource = '';
            links.forEach(l => {{
                if (selectedIds.has(l.source) && selectedIds.has(l.target)) {{
                    relSource += `    ${{l.source}} ||--o{{ ${{l.target}} : "${{l.label}}"\n`;
                }}
            }});
            pre.textContent += relSource;

            // 更新进度
            const progressEl = document.getElementById('render-progress');
            if (progressEl) progressEl.textContent = '⏳ Mermaid 布局计算中...';

            // 延迟一帧让浏览器更新 DOM，然后启动 Mermaid
            requestAnimationFrame(() => {{
                mermaid.run({{
                    nodes: [pre]
                }}).then(() => {{
                    showProgress(false);
                    isRendering = false;
                    if (renderQueue.length > 0) {{
                        renderQueue = [];
                        startRender();
                    }}
                }}).catch(err => {{
                    console.error('Mermaid render error:', err);
                    showProgress(false);
                    isRendering = false;
                }});
            }});
        }}

        // 初始化 Mermaid - 优化配置 (maxTextSize 设为大数防止大图报错，0 会被 JS 当作 falsy 回退默认值)
        mermaid.initialize({{
            startOnLoad: false,
            theme: 'default',
            maxTextSize: 99999999,
            securityLevel: 'loose',
            er: {{
                useMaxWidth: false,
                layoutDirection: 'TB',
                minEntityHeight: 20,
                entityPadding: 10
            }}
        }});

        // ── 缩放与平移 ──
        let scale = 1, pointX = 0, pointY = 0, startX = 0, startY = 0, isDragging = false;
        const viewport = document.getElementById('viewport');
        const wrapper = document.getElementById('diagram-wrapper');
        function updateTransform() {{ wrapper.style.transform = `translate(${{pointX}}px, ${{pointY}}px) scale(${{scale}})`; }}
        viewport.onmousedown = e => {{ if(e.button===0) {{ startX = e.clientX - pointX; startY = e.clientY - pointY; isDragging = true; }} }};
        window.onmouseup = () => isDragging = false;
        window.onmousemove = e => {{ if (isDragging) {{ pointX = e.clientX - startX; pointY = e.clientY - startY; updateTransform(); }} }};
        viewport.onwheel = e => {{ e.preventDefault(); const factor = 0.1; scale = e.deltaY > 0 ? Math.max(0.1, scale-factor) : Math.min(5, scale+factor); updateTransform(); }};
        function resetView() {{ scale = 0.6; pointX = viewport.offsetWidth / 4; pointY = 100; updateTransform(); }}

        // 启动
        initList();
        startRender();
        setTimeout(resetView, 1000);
    </script>
</body>
</html>
"""
        with open(file_path, "w", encoding='utf-8') as f:
            f.write(html_template.replace('<MERMAID_SCRIPT>', mermaid_script))

    @staticmethod
    def export_table_structure_to_pdf(structures, file_path):
        """将表结构导出为 PDF 文件 (包含目录)"""
        from fpdf import FPDF
        import warnings
        
        # 忽略 fontTools 的子集化警告
        warnings.filterwarnings("ignore", message=".*MERG NOT subset.*")
        
        class PDF(FPDF):
            def header(self):
                if self.page_no() > 1: # 目录页之后显示页眉
                    self.set_font('chinese', 'B', 15)
                    self.cell(0, 10, '数据库表结构报告', 0, 1, 'C')
                    self.ln(5)

            def footer(self):
                self.set_y(-15)
                self.set_font('chinese', 'I', 8)
                self.cell(0, 10, f'第 {self.page_no()} 页', 0, 0, 'C')

        def render_toc(pdf, outline):
            """渲染目录页的具体逻辑 (修复上下一半显示不全的问题)"""
            pdf.set_font("chinese", "B", 16)
            pdf.cell(0, 25, "目录", ln=1, align="C")
            pdf.set_font("chinese", "", 12)
            
            # 增加行高，给中文字体留足上下空间，防止被切掉一半
            row_height = 12 
            page_num_width = 15
            
            for entry in outline:
                link = pdf.add_link()
                pdf.set_link(link, page=entry.page_number)
                
                text = f"{entry.name}"
                text_width = pdf.get_string_width(text)
                max_text_width = pdf.epw - page_num_width - 15
                
                if text_width > max_text_width:
                    while pdf.get_string_width(text + "...") > max_text_width:
                        text = text[:-1]
                    text = text + "..."
                    text_width = pdf.get_string_width(text)

                # 1. 绘制表名
                pdf.cell(text_width, row_height, text, link=link)
                
                # 2. 虚线填充
                dot_width = pdf.epw - text_width - page_num_width
                dot_count = int(dot_width / pdf.get_string_width(".")) - 2
                if dot_count > 0:
                    pdf.cell(dot_width, row_height, "." * dot_count, align="R")
                else:
                    pdf.set_x(pdf.get_x() + dot_width)
                
                # 3. 页码
                pdf.cell(page_num_width, row_height, str(entry.page_number), ln=1, align="R")

        pdf = PDF()
        
        # 加载字体
        font_paths = ["C:\\Windows\\Fonts\\simhei.ttf", "C:\\Windows\\Fonts\\msyh.ttc", "C:\\Windows\\Fonts\\simsun.ttc"]
        for path in font_paths:
            if os.path.exists(path):
                pdf.add_font('chinese', '', path)
                pdf.add_font('chinese', 'B', path)
                pdf.add_font('chinese', 'I', path)
                break
        
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # --- 核心修复：精确计算目录所需页数 ---
        # 1. 计算可用空间
        import math
        usable_height = 297 - 15 - 15 # A4 高度 297mm，上下边距 15mm
        title_height = 25 # 目录标题占用高度
        row_height = 12 # 目录每行高度
        
        num_tables = len(structures)
        # 第一页能放多少行
        first_page_rows = math.floor((usable_height - title_height) / row_height)
        # 后续页能放多少行
        other_page_rows = math.floor(usable_height / row_height)
        
        if num_tables <= first_page_rows:
            expected_toc_pages = 1
        else:
            expected_toc_pages = 1 + math.ceil((num_tables - first_page_rows) / other_page_rows)
        
        # 插入目录占位符
        for _ in range(expected_toc_pages):
            pdf.add_page()
        pdf.insert_toc_placeholder(render_toc, pages=expected_toc_pages)

        # 2. 正常渲染表结构
        sorted_tables = sorted(structures.keys())
        for table_name in sorted_tables:
            # 开启新章节，这会自动被记录到 outline 中用于生成目录
            pdf.start_section(table_name)
            
            info = structures[table_name]
            pdf.set_font('chinese', 'B', 12)
            pdf.set_fill_color(200, 220, 255)
            pdf.cell(0, 10, f"表名: {table_name}", 0, 1, 'L', fill=True)
            
            pdf.set_font('chinese', '', 10)
            pdf.multi_cell(0, 8, f"备注: {info['comment'] or '无'}")
            pdf.ln(2)

            # 表格绘制
            headers = ["字段", "类型", "允许空", "键", "默认", "备注"]
            widths = [35, 45, 15, 10, 25, 60]
            with pdf.table(col_widths=widths, borders_layout="ALL", line_height=6, text_align="LEFT") as table:
                pdf.set_font('chinese', 'B', 9)
                pdf.set_fill_color(240, 240, 240)
                header_row = table.row()
                for h in headers: header_row.cell(h)
                
                pdf.set_font('chinese', '', 8)
                for col in info['columns']:
                    row_data = [col['Field'], col['Type'], col['Null'], col['Key'], str(col['Default']) if col['Default'] is not None else '', col['Comment']]
                    row = table.row()
                    for val in row_data: row.cell(str(val))
            pdf.ln(10)

        pdf.output(file_path)

    @staticmethod
    def export_table_structure_to_markdown(structures, file_path):
        """将表结构导出为 Markdown 文件"""
        lines = ["# 数据库表结构报告\n"]
        
        sorted_tables = sorted(structures.keys())
        
        # 生成详情
        for table_name in sorted_tables:
            info = structures[table_name]
            # 使用 ## 级别标题，确保出现在编辑器的右侧导航/大纲列表中
            lines.append(f"## {table_name}")
            lines.append(f"**表注释:** {info['comment'] or '无'}\n")
            
            headers = ["字段", "类型", "字符集", "允许空", "键", "默认值", "额外", "备注"]
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            
            for col in info['columns']:
                row = [
                    col['Field'], col['Type'], col.get('Collation', '') or '', col['Null'],
                    col['Key'], str(col['Default']) if col['Default'] is not None else '',
                    col['Extra'], col['Comment']
                ]
                lines.append("| " + " | ".join(map(str, row)) + " |")
            lines.append("\n---\n")
            
        with open(file_path, "w", encoding='utf-8') as f:
            f.write("\n".join(lines))

    @staticmethod
    def export_table_structure_to_excel_single_sheet(structures, file_path):
        """将所有表结构导出到同一个 Excel Sheet 中"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = Exporter._sanitize_sheet_name("所有表结构")
        
        # 定义样式
        table_header_fill = PatternFill(start_color="007BBB", end_color="007BBB", fill_type="solid")
        table_header_font = Font(bold=True, color="FFFFFF", size=12)
        
        column_header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        column_header_font = Font(bold=True)

        current_row = 1
        sorted_tables = sorted(structures.keys())
        
        for table_name in sorted_tables:
            info = structures[table_name]
            
            # 1. 写入大标题 (表名 + 注释)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            safe_table_title = Exporter._clean_value(f"数据表: {table_name}  (备注: {info['comment'] or '无'})")
            cell = ws.cell(row=current_row, column=1, value=safe_table_title)
            cell.fill = table_header_fill
            cell.font = table_header_font
            current_row += 1
            
            # 2. 写入列头
            headers = ["字段", "类型", "字符集", "允许空", "键", "默认值", "额外", "权限", "备注"]
            ws.append([Exporter._clean_value(h) for h in headers])
            for col_idx in range(1, 10):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = column_header_fill
                cell.font = column_header_font
            current_row += 1
            
            # 3. 写入字段数据
            for col in info['columns']:
                row_data = [
                    col['Field'], col['Type'], col.get('Collation', '') or '', col['Null'],
                    col['Key'], col['Default'], col['Extra'], col.get('Privileges', '') or '',
                    col['Comment']
                ]
                ws.append([Exporter._clean_value(v) for v in row_data])
                current_row += 1
            
            # 4. 插入两个空行作为表之间的间隔
            ws.append([])
            ws.append([])
            current_row += 2

        # 调整列宽
        for col_idx in range(1, 10):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20 # 设置固定宽度后微调

        wb.save(file_path)

    @staticmethod
    def export_table_structure_to_excel(structures, file_path):
        """将表结构导出到 Excel (多 Sheet)"""
        wb = openpyxl.Workbook()
        # 删除默认 Sheet
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)
        
        for table_name, info in structures.items():
            # 使用统一清洗逻辑，解决非法字符和超长问题
            sheet_name = Exporter._sanitize_sheet_name(table_name)
            ws = wb.create_sheet(title=sheet_name)
            
            # 写入表头信息
            header_info = [f"表名: {table_name}", f"注释: {info['comment']}"]
            ws.append([Exporter._clean_value(v) for v in header_info])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.append([]) # 空行
            
            # 写入列头
            headers = ["字段", "类型", "字符集", "允许空", "键", "默认值", "额外", "权限", "备注"]
            ws.append([Exporter._clean_value(h) for h in headers])
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
            
            # 写入列数据
            for col in info['columns']:
                row_data = [
                    col['Field'], col['Type'], col.get('Collation', '') or '', col['Null'],
                    col['Key'], col['Default'], col['Extra'], col.get('Privileges', '') or '',
                    col['Comment']
                ]
                ws.append([Exporter._clean_value(v) for v in row_data])
            
            # 调整列宽
            for col_idx in range(1, ws.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = 0
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(file_path)

    @staticmethod
    def export_connections_to_excel(connections, file_path, include_sensitive=False):
        """将连接信息导出到 Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = Exporter._sanitize_sheet_name("Connections")
        
        headers = [
            "名称", "主机", "端口", "用户", "数据库", "启用SSH", 
            "SSH主机", "SSH端口", "SSH用户", "SSH认证方式", "SSH本地端口", "SSH远程主机", "SSH远程端口"
        ]
        if include_sensitive:
            headers.extend(["密码", "SSH密码", "SSH私钥路径", "SSH私钥密码"])
        
        ws.append([Exporter._clean_value(h) for h in headers])
        
        for conn in connections:
            row = [
                conn['name'], conn['host'], conn['port'], conn['user'], conn['database'],
                "是" if conn['ssh_enabled'] else "否",
                conn['ssh_host'], conn['ssh_port'], conn['ssh_user'], conn['ssh_auth_type'],
                conn['ssh_local_port'], conn['ssh_remote_host'], conn['ssh_remote_port']
            ]
            if include_sensitive:
                row.extend([
                    conn.get('password', ''), conn.get('ssh_password', ''),
                    conn.get('ssh_key_path', ''), conn.get('ssh_key_phrase', '')
                ])
            ws.append([Exporter._clean_value(v) for v in row])
            
        wb.save(file_path)

    @staticmethod
    def export_to_navicat(connections, file_path):
        """导出所有连接为 Navicat .ncx 文件 (支持 MySQL 和 PostgreSQL)"""
        cipher = NavicatCipher()
        
        # 创建根节点 <Connections Ver="1.5">
        root = ET.Element("Connections")
        root.set("Ver", "1.5")
        
        for conn_data in connections:
            db_type = conn_data.get('db_type', 'MySQL')
            navicat_type = "MYSQL" if db_type == "MySQL" else "POSTGRESQL"
            
            # 创建 Connection 节点，所有属性平铺
            conn_el = ET.SubElement(root, "Connection")
            
            # 基础必填属性
            conn_el.set("ConnectionName", conn_data['name'])
            conn_el.set("ConnType", navicat_type)
            conn_el.set("Host", conn_data['host'])
            conn_el.set("Port", str(conn_data['port']))
            conn_el.set("UserName", conn_data['user'])
            
            # 补充模板中的默认属性
            conn_el.set("ProjectUUID", "")
            conn_el.set("ServiceProvider", "Default")
            conn_el.set("SettingsSavePath", "")
            conn_el.set("SessionLimit", "0")
            conn_el.set("InitialSessionQueries", "")
            conn_el.set("ClientDriverVersion", "Default")
            conn_el.set("ClientCharacterSet", "")
            conn_el.set("ClientEncoding", "65001")
            conn_el.set("Keepalive", "false")
            conn_el.set("UseConnectionTimeout", "true")
            conn_el.set("ConnectionTimeoutSeconds", "30")
            conn_el.set("UseReadTimeout", "false")
            conn_el.set("UseWriteTimeout", "true")
            conn_el.set("WriteTimeoutSeconds", "30")
            conn_el.set("Encoding", "65001")
            conn_el.set("MySQLCharacterSet", "true")
            conn_el.set("Compression", "false")
            conn_el.set("AutoConnect", "false")
            conn_el.set("NamedPipe", "false")
            conn_el.set("UseAdvanced", "false")
            conn_el.set("SSL", "false")
            conn_el.set("HTTP", "false")
            conn_el.set("Compatibility", "false")
            conn_el.set("Remarks", "")

            # 密码处理 (使用 Navicat 12+ 专用 AES 算法)
            if conn_data.get('password'):
                 conn_el.set("Password", cipher.encrypt(conn_data['password']))

            # SSH 逻辑处理
            if conn_data.get('ssh_enabled'):
                conn_el.set("SSH", "true")
                conn_el.set("SSHHost", conn_data['ssh_host'])
                conn_el.set("SSHPort", str(conn_data['ssh_port'] or 22))
                conn_el.set("SSHUserName", conn_data['ssh_user'])
                
                if conn_data['ssh_auth_type'] == 'password':
                    conn_el.set("SSHAuthType", "0") # 密码
                    if conn_data.get('ssh_password'):
                        conn_el.set("SSHPassword", cipher.encrypt(conn_data['ssh_password']))
                else:
                    conn_el.set("SSHAuthType", "1") # 公钥
                    conn_el.set("SSHKeyFile", conn_data.get('ssh_key_path', ''))
                    if conn_data.get('ssh_key_phrase'):
                        conn_el.set("SSHPassphrase", cipher.encrypt(conn_data['ssh_key_phrase']))
            else:
                conn_el.set("SSH", "false")

        # 格式化并保存
        xml_str = ET.tostring(root, encoding='utf-8')
        reparsed = minidom.parseString(xml_str)
        pretty_xml = reparsed.toprettyxml(indent="\t", encoding="UTF-8").decode('utf-8')
        
        with open(file_path, "w", encoding='utf-8') as f:
            f.write(pretty_xml)
