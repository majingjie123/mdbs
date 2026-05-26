"""数据导出 API —— 表结构 / ER 图 / 数据 / 连接配置"""

import os
import json
import tempfile
import shutil
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse

from dependencies import get_db_storage, get_db_ops
from schemas import (
    ExportStructureRequest,
    ExportERRequest,
    ExportNavicatRequest,
    ExportDataRequest,
    MessageResponse,
)
from core.db_operations import DBOperations
from models.db_storage import DBStorage
from core.exporter import Exporter

router = APIRouter(prefix="/api/export", tags=["导出"])

EXPORT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "sync_logs",
    "exports",
)


def _ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def _get_conn_data(conn_id: int, storage: DBStorage):
    conn = storage.get_connection(conn_id)
    if not conn:
        raise HTTPException(status_code=404, detail="连接不存在")
    return conn


def _collect_structures(conn_data: dict, ops: DBOperations, db: str, schema: str, tables: list[str] = None):
    """收集表结构数据"""
    all_structures = ops.get_table_structure(conn_data, database=db or None, schema=schema or None)
    if tables:
        filtered = {}
        for t in tables:
            if t in all_structures:
                filtered[t] = all_structures[t]
        return filtered
    return all_structures


@router.post("/structure")
def export_structure(body: ExportStructureRequest, storage: DBStorage = Depends(get_db_storage), ops: DBOperations = Depends(get_db_ops)):
    """导出表结构"""
    conn_data = _get_conn_data(body.conn_id, storage)
    structures = _collect_structures(conn_data, ops, body.database, body.schema_name, body.tables)
    if not structures:
        raise HTTPException(status_code=404, detail="未找到表结构数据")

    _ensure_export_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmt = body.format.lower()

    ext_map = {"excel": "xlsx", "pdf": "pdf", "html": "html", "markdown": "md"}
    ext = ext_map.get(fmt, "xlsx")
    file_path = os.path.join(EXPORT_DIR, f"table_structure_{ts}.{ext}")

    if fmt == "excel":
        Exporter.export_table_structure_to_excel(structures, file_path)
    elif fmt == "pdf":
        Exporter.export_table_structure_to_pdf(structures, file_path)
    elif fmt == "html":
        Exporter.export_table_structure_to_html(structures, file_path)
    elif fmt == "markdown":
        Exporter.export_table_structure_to_markdown(structures, file_path)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的格式: {fmt}")

    return FileResponse(file_path, filename=os.path.basename(file_path))


@router.post("/er")
def export_er(body: ExportERRequest, storage: DBStorage = Depends(get_db_storage), ops: DBOperations = Depends(get_db_ops)):
    """导出 ER 图"""
    conn_data = _get_conn_data(body.conn_id, storage)
    structures = _collect_structures(conn_data, ops, body.database, body.schema_name)
    if not structures:
        raise HTTPException(status_code=404, detail="未找到表结构数据")

    relations = []
    if body.include_relations:
        rels = ops.get_relations(conn_data, database=body.database or None)
        if rels:
            relations = rels

    _ensure_export_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmt = body.format.lower()

    ext_map = {"excel": "xlsx", "pdf": "pdf", "html": "html", "markdown": "md"}
    ext = ext_map.get(fmt, "html")
    file_path = os.path.join(EXPORT_DIR, f"er_diagram_{ts}.{ext}")

    if fmt == "excel":
        Exporter.export_er_to_excel(structures, relations, file_path)
    elif fmt == "pdf":
        Exporter.export_er_to_pdf(structures, relations, file_path)
    elif fmt == "html":
        Exporter.export_er_to_html(structures, relations, file_path)
    elif fmt == "markdown":
        Exporter.export_er_to_markdown(structures, relations, file_path)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的格式: {fmt}")

    return FileResponse(file_path, filename=os.path.basename(file_path))


@router.post("/navicat")
def export_navicat(body: ExportNavicatRequest, storage: DBStorage = Depends(get_db_storage)):
    """导出 Navicat 连接配置 (.ncx)"""
    connections = []
    for cid in body.conn_ids:
        conn = storage.get_connection(cid)
        if conn:
            connections.append(conn)

    if not connections:
        raise HTTPException(status_code=404, detail="未找到连接配置")

    _ensure_export_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(EXPORT_DIR, f"navicat_connections_{ts}.ncx")
    Exporter.export_to_navicat(connections, file_path)

    return FileResponse(file_path, filename=os.path.basename(file_path))


def clean_old_temp_files(days_limit=7, size_limit_mb=200):
    """自动清理旧的同步日志和导出临时文件"""
    import time
    from pathlib import Path
    
    base_dir = Path(EXPORT_DIR).parent  # 获取 sync_logs 目录
    if not base_dir.exists():
        return
        
    now = time.time()
    cutoff = now - (days_limit * 86400)
    
    all_files = []
    # 1. 递归扫描并按时间清理
    for p in base_dir.glob("**/*"):
        if not p.is_file():
            continue
        if p.stat().st_mtime < cutoff:
            try:
                p.unlink()
            except:
                pass
        else:
            all_files.append(p)
            
    # 2. 按总体积限制清理
    total_size = sum(f.stat().st_size for f in all_files)
    if total_size > size_limit_mb * 1024 * 1024:
        all_files.sort(key=lambda x: x.stat().st_mtime)
        accumulated_size = total_size
        limit = size_limit_mb * 1024 * 1024
        for f in all_files:
            if accumulated_size <= limit:
                break
            try:
                f_size = f.stat().st_size
                f.unlink()
                accumulated_size -= f_size
            except:
                pass


@router.post("/data")
def export_data(body: ExportDataRequest, storage: DBStorage = Depends(get_db_storage), ops: DBOperations = Depends(get_db_ops)):
    """导出表数据（流式分批导出，防止内存溢出）"""
    conn_data = _get_conn_data(body.conn_id, storage)
    db_type = conn_data.get("db_type", "MySQL")
    quote = "`" if db_type == "MySQL" else '"'
    sql = f"SELECT * FROM {quote}{body.table}{quote}"

    # 1. 建立底层物理连接
    conn = ops.get_connection(conn_data, database=body.database or None)
    cols = []

    # 2. 提取表头字段名（卫语句防御）
    try:
        if db_type == "MySQL":
            with conn.cursor() as cursor:
                cursor.execute(f"SELECT * FROM `{body.table}` LIMIT 0")
                cols = [col[0] for col in cursor.description]
        else:
            conn.run("BEGIN")
            conn.run(f'DECLARE cols_cursor CURSOR FOR SELECT * FROM "{body.table}" LIMIT 0')
            conn.run('FETCH 0 FROM cols_cursor')
            cols = [col['name'] for col in conn.columns] if hasattr(conn, 'columns') else []
            conn.run('CLOSE cols_cursor')
            conn.run('COMMIT')
    except Exception as e:
        if db_type != "MySQL":
            try: conn.run("ROLLBACK")
            except: pass
        if hasattr(conn, 'close'): conn.close()
        raise HTTPException(status_code=500, detail=f"获取表头结构失败: {e}")

    # 3. 开始写出到本地文件
    _ensure_export_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmt = body.format.lower()
    batch_size = 1000

    if fmt == "csv":
        file_path = os.path.join(EXPORT_DIR, f"data_{body.table}_{ts}.csv")
        try:
            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                import csv
                writer = csv.writer(f, quoting=csv.QUOTE_ALL)
                if cols:
                    writer.writerow(cols)

                # 分批拉取并增量写入 CSV
                if db_type == "MySQL":
                    import pymysql.cursors
                    with conn.cursor(pymysql.cursors.SSCursor) as cursor:
                        cursor.execute(sql)
                        while True:
                            rows = cursor.fetchmany(batch_size)
                            if not rows:
                                break
                            writer.writerows(rows)
                else:
                    conn.run("BEGIN")
                    conn.run(f'DECLARE export_cursor CURSOR FOR SELECT * FROM "{body.table}"')
                    while True:
                        rows = conn.run(f"FETCH {batch_size} FROM export_cursor")
                        if not rows or len(rows) == 0:
                            break
                        writer.writerows(rows)
                    conn.run("CLOSE export_cursor")
                    conn.run("COMMIT")
        except Exception as e:
            if db_type != "MySQL":
                try: conn.run("ROLLBACK")
                except: pass
            raise HTTPException(status_code=500, detail=f"写入 CSV 失败: {e}")
        finally:
            if hasattr(conn, 'close'): conn.close()

    elif fmt == "excel":
        file_path = os.path.join(EXPORT_DIR, f"data_{body.table}_{ts}.xlsx")
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # 使用 openpyxl 的 write_only 模式，实现增量流式生成 Excel
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title=body.table[:31])
            
            if cols:
                ws.append(cols)

            col_widths = {i+1: len(str(c)) for i, c in enumerate(cols)} if cols else {}
            sample_lines = 0

            # 分批拉取并追加写入 Excel
            if db_type == "MySQL":
                import pymysql.cursors
                with conn.cursor(pymysql.cursors.SSCursor) as cursor:
                    cursor.execute(sql)
                    while True:
                        rows = cursor.fetchmany(batch_size)
                        if not rows:
                            break
                        for row in rows:
                            cleaned_row = [v if v is not None else "" for v in row]
                            ws.append(cleaned_row)
                            if sample_lines < 1000:
                                sample_lines += 1
                                for idx, val in enumerate(cleaned_row):
                                    col_widths[idx+1] = max(col_widths.get(idx+1, 10), len(str(val)))
            else:
                conn.run("BEGIN")
                conn.run(f'DECLARE export_cursor CURSOR FOR SELECT * FROM "{body.table}"')
                while True:
                    rows = conn.run(f"FETCH {batch_size} FROM export_cursor")
                    if not rows or len(rows) == 0:
                        break
                    for row in rows:
                        cleaned_row = [v if v is not None else "" for v in row]
                        ws.append(cleaned_row)
                        if sample_lines < 1000:
                            sample_lines += 1
                            for idx, val in enumerate(cleaned_row):
                                col_widths[idx+1] = max(col_widths.get(idx+1, 10), len(str(val)))
                conn.run("CLOSE export_cursor")
                conn.run("COMMIT")

            # 在只写模式下配置自适应列宽（采样前1000行）
            for col_idx, width in col_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(width + 2, 50)

            wb.save(file_path)
        except Exception as e:
            if db_type != "MySQL":
                try: conn.run("ROLLBACK")
                except: pass
            raise HTTPException(status_code=500, detail=f"写入 Excel 失败: {e}")
        finally:
            if hasattr(conn, 'close'): conn.close()
    else:
        if hasattr(conn, 'close'): conn.close()
        raise HTTPException(status_code=400, detail=f"不支持的格式: {fmt}")

    # 导出完成后，异步触发日志与导出文件轮转清理
    try:
        clean_old_temp_files()
    except:
        pass

    return FileResponse(file_path, filename=os.path.basename(file_path))