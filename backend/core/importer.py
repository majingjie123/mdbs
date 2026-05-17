"""
CSV / Excel 数据导入模块
遵循 core/exporter.py 的 @staticmethod 风格，与 UI 解耦。
"""
import csv
import re
import io
from datetime import datetime

# 检测 datetime 类型值的正则（常用格式）
_DATETIME_PAT = re.compile(
    r'^\d{4}-\d{2}-\d{2}(\s+\d{2}:\d{2}(:\d{2})?)?$|'
    r'^\d{2}/\d{2}/\d{4}(\s+\d{2}:\d{2}(:\d{2})?)?$'
)

class Importer:
    # ----------------------------------------------------------------
    # 文件解析
    # ----------------------------------------------------------------

    @staticmethod
    def _try_parse_value(val):
        """尝试将字符串解析为 Python 类型（int → float → datetime → str）"""
        if val is None or val.strip() == '':
            return None
        s = val.strip()
        # int
        try:
            return int(s)
        except ValueError:
            pass
        # float
        try:
            return float(s)
        except ValueError:
            pass
        # datetime
        if _DATETIME_PAT.match(s):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                         '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y'):
                try:
                    return datetime.strptime(s, fmt)
                except ValueError:
                    continue
        return s

    @staticmethod
    def parse_csv(file_path, delimiter=',', has_header=True, encoding=None):
        """
        解析 CSV 文件。
        返回 (columns, rows_display, rows_raw)
        - columns: 列名列表
        - rows_display: 所有单元格为 str 的行列表，用于预览
        - rows_raw: 所有单元格为 Python 原生类型的行列表，用于导入
        """
        detected_enc = encoding or Importer._detect_encoding(file_path)
        rows_raw = []
        with open(file_path, 'r', encoding=detected_enc, errors='replace') as f:
            reader = csv.reader(f, delimiter=delimiter)
            raw = list(reader)

        if not raw:
            return [], [], []

        if has_header:
            columns = [c.strip() or f"col{i}" for i, c in enumerate(raw[0])]
            data_start = 1
        else:
            columns = [f"col{i+1}" for i in range(len(raw[0]))]
            data_start = 0

        for row in raw[data_start:]:
            if not row or all(c.strip() == '' for c in row):
                continue
            rows_raw.append([Importer._try_parse_value(c) for c in row])

        # rows_display: 所有值转 str
        rows_display = [[str(v) if v is not None else '' for v in row] for row in rows_raw]

        # 补齐列数
        ncols = len(columns)
        for r in rows_display:
            while len(r) < ncols:
                r.append('')
        for r in rows_raw:
            while len(r) < ncols:
                r.append(None)

        return columns, rows_display, rows_raw

    @staticmethod
    def parse_excel(file_path, sheet_name=None, has_header=True):
        """
        解析 Excel (.xlsx) 文件。
        返回 (columns, rows_display, rows_raw)
        """
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        raw_rows = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))

        wb.close()

        if not raw_rows:
            return [], [], []

        if has_header:
            columns = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(raw_rows[0])]
            data_start = 1
        else:
            columns = [f"col{i+1}" for i in range(len(raw_rows[0]))]
            data_start = 0

        rows_raw = []
        for row in raw_rows[data_start:]:
            if all(v is None or (isinstance(v, str) and v.strip() == '') for v in row):
                continue
            rows_raw.append([v if v is not None else None for v in row])

        rows_display = [[str(v) if v is not None else '' for v in row] for row in rows_raw]

        ncols = len(columns)
        for r in rows_display:
            while len(r) < ncols:
                r.append('')
        for r in rows_raw:
            while len(r) < ncols:
                r.append(None)

        return columns, rows_display, rows_raw

    @staticmethod
    def _detect_encoding(file_path, sample_size=8192):
        """检测文件编码（BOM 优先，否则 utf-8 回退）"""
        with open(file_path, 'rb') as f:
            raw = f.read(sample_size)
        if raw[:3] == b'\xef\xbb\xbf':
            return 'utf-8-sig'
        if raw[:2] == b'\xff\xfe':
            return 'utf-16-le'
        if raw[:2] == b'\xfe\xff':
            return 'utf-16-be'
        try:
            raw.decode('utf-8')
            return 'utf-8'
        except UnicodeDecodeError:
            try:
                raw.decode('gbk')
                return 'gbk'
            except UnicodeDecodeError:
                return 'utf-8'

    # ----------------------------------------------------------------
    # 类型推断 & CREATE TABLE
    # ----------------------------------------------------------------

    @staticmethod
    def _infer_type(values):
        """根据列值集合推断 SQL 列类型"""
        has_int = False
        has_float = False
        has_str = False
        has_date = False
        max_len = 0

        for v in values:
            if v is None:
                continue
            if isinstance(v, int):
                if not has_float:
                    has_int = True
            elif isinstance(v, float):
                has_float = True
                has_int = False
            elif isinstance(v, datetime):
                has_date = True
            elif isinstance(v, str):
                has_str = True
                max_len = max(max_len, len(v))
            else:
                has_str = True
                s = str(v)
                max_len = max(max_len, len(s))

        if has_date:
            return 'DATETIME'
        if has_float:
            return 'DECIMAL(18, 6)'
        if has_int and not has_str:
            return 'BIGINT'
        # 字符串
        if max_len < 1:
            max_len = 1
        if max_len > 1000:
            return 'TEXT'
        return f'VARCHAR({max(max_len, 1) * 2})'

    @staticmethod
    def generate_create_sql(table_name, columns, rows_raw, db_type='MySQL'):
        """根据数据推断列类型并生成 CREATE TABLE 语句"""
        col_types = []
        for i, col in enumerate(columns):
            vals = [r[i] for r in rows_raw if i < len(r)]
            col_types.append(Importer._infer_type(vals))

        cols_def = []
        for col, typ in zip(columns, col_types):
            safe_col = f"`{col}" if db_type == 'MySQL' else f'"{col}"'
            cols_def.append(f"  {safe_col} {typ}")

        if db_type == 'MySQL':
            return f"CREATE TABLE `{table_name}` (\n" + ",\n".join(cols_def) + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;"
        else:
            return f'CREATE TABLE "{table_name}" (\n' + ",\n".join(cols_def) + "\n);"

    # ----------------------------------------------------------------
    # 数据导入（写入数据库）
    # ----------------------------------------------------------------

    @staticmethod
    def import_to_table(db_ops, conn_data, table_name, columns, rows_raw,
                        mode='append', database=None, schema=None,
                        cancel_event=None, progress_callback=None,
                        batch_size=500):
        """
        将数据导入数据库表。

        参数:
            mode: 'append' — INSERT 追加
                  'replace' — TRUNCATE + INSERT
                  'create'  — CREATE TABLE + INSERT

        返回 (success, affected_rows_or_error_msg)
        """
        db_type = conn_data.get('db_type', 'MySQL')

        # 1. 获取连接
        try:
            conn = db_ops.get_connection(conn_data, database=database)
        except Exception as e:
            return False, f"无法连接数据库: {e}"

        try:
            # 2. 前置处理
            if mode == 'replace':
                if db_type == 'MySQL':
                    with conn.cursor() as cur:
                        cur.execute(f"DELETE FROM `{table_name}`")
                    conn.commit()
                else:
                    conn.run(f'DELETE FROM "{table_name}"')
                    conn.run("COMMIT")

            elif mode == 'create':
                create_sql = Importer.generate_create_sql(table_name, columns, rows_raw, db_type)
                if db_type == 'MySQL':
                    with conn.cursor() as cur:
                        cur.execute(create_sql)
                    conn.commit()
                else:
                    conn.run(create_sql)
                    conn.run("COMMIT")

            # 3. 批量 INSERT（多值语句，大幅减少网络往返）
            total = len(rows_raw)
            num_cols = len(columns)
            safe_table = f"`{table_name}`" if db_type == 'MySQL' else f'"{table_name}"'
            safe_cols = ', '.join([f"`{c}`" if db_type == 'MySQL' else f'"{c}"' for c in columns])

            def _fmt_row(vals):
                """将 datetime 转为字符串"""
                return tuple(
                    v.strftime('%Y-%m-%d %H:%M:%S') if isinstance(v, datetime) else v
                    for v in vals
                )

            affected = 0
            if db_type == 'MySQL':
                with conn.cursor() as cur:
                    if total > 1:
                        cur.execute("SET autocommit=0")
                    # 按 batch_size 分批构造多值 INSERT
                    for batch_start in range(0, total, batch_size):
                        if cancel_event and cancel_event.is_set():
                            raise Exception("导入已取消")
                        batch = rows_raw[batch_start:batch_start + batch_size]
                        batch_params = []
                        for row in batch:
                            batch_params.extend(_fmt_row(row))
                        # 构造多值占位符: (%s,%s),(%s,%s),...
                        row_placeholders = ','.join(
                            '(' + ','.join(['%s'] * num_cols) + ')'
                            for _ in range(len(batch))
                        )
                        sql = f"INSERT INTO {safe_table} ({safe_cols}) VALUES {row_placeholders}"
                        cur.execute(sql, tuple(batch_params))
                        affected += len(batch)
                        if progress_callback:
                            pct = min(100, int(affected / total * 100))
                            progress_callback(pct, f"正在导入第 {affected}/{total} 行...")
                    if total > 1:
                        conn.commit()
            else:
                # PostgreSQL 批量多值 INSERT
                try:
                    conn.run("BEGIN")
                    for batch_start in range(0, total, batch_size):
                        if cancel_event and cancel_event.is_set():
                            raise Exception("导入已取消")
                        batch = rows_raw[batch_start:batch_start + batch_size]
                        batch_params = []
                        for row in batch:
                            batch_params.extend(_fmt_row(row))
                        # 构造多值占位符: ($1,$2),($3,$4),...
                        param_idx = 1
                        row_placeholders_list = []
                        for _ in range(len(batch)):
                            row_placeholders_list.append(
                                '(' + ','.join(f'${param_idx + c}' for c in range(num_cols)) + ')'
                            )
                            param_idx += num_cols
                        row_placeholders = ','.join(row_placeholders_list)
                        sql = f"INSERT INTO {safe_table} ({safe_cols}) VALUES {row_placeholders}"
                        conn.run(sql, *batch_params)
                        affected += len(batch)
                        if progress_callback:
                            pct = min(100, int(affected / total * 100))
                            progress_callback(pct, f"正在导入第 {affected}/{total} 行...")
                    conn.run("COMMIT")
                except Exception:
                    conn.run("ROLLBACK")
                    raise

            return True, affected

        except Exception as e:
            try:
                if db_type == 'MySQL':
                    conn.rollback()
            except Exception:
                pass
            return False, str(e)
        finally:
            try:
                conn.close()
            except Exception:
                pass

    @staticmethod
    def get_excel_sheet_names(file_path):
        """获取 Excel 文件的工作表名列表"""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
